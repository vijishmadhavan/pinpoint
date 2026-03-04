"""
Pinpoint — FastAPI server (bridge between Node.js and Python)

Endpoints:
  GET  /ping                            → health check
  GET  /search?q=...                    → FTS5 search (with file_type/folder filters)
  GET  /status                          → indexing statistics
  POST /index                           → index a folder
  GET  /document/{id}                   → get full document text by ID
  GET  /list_files?folder=              → list folder contents (sort/filter)
  GET  /file_info?path=                 → file metadata (size, dates, indexed status)
  POST /calculate                       → safe math evaluation
  POST /read_excel                      → read Excel cells/ranges via openpyxl
  POST /move_file                       → move/copy/rename files
  POST /create_folder                   → create directories
  POST /delete_file                     → delete a file (safety checks)
  POST /read_file                       → read actual file from disk (images→base64, docs→text)
  POST /conversation                    → save a conversation message
  GET  /conversation/history            → load conversation history
  POST /conversation/reset              → reset a conversation session
  GET  /conversation/search             → search past conversations
  POST /detect-faces                    → detect faces in an image
  POST /crop-face                       → crop a specific face from an image
  POST /find-person                     → find matching faces in a folder
  POST /find-person-by-face             → find matches using a specific face index
  POST /count-faces                     → count faces with age/gender summary
  POST /compare-faces                   → compare two faces (same person check)
  POST /remember-face                   → save a face for future recognition
  POST /forget-face                     → delete saved face data for a person
  POST /recognize-faces                 → recognize known faces in an image
  POST /ocr                             → OCR text extraction (Tesseract)

  POST /analyze-data                    → Pandas data analysis on CSV/Excel
  POST /index-file                      → Index a single file on demand
  POST /write-file                      → Create/write text files
  POST /generate-excel                  → Create Excel from data
  POST /generate-chart                  → Create chart image (matplotlib)
  POST /merge-pdf                       → Combine multiple PDFs
  POST /split-pdf                       → Extract pages from PDF
  POST /resize-image                    → Resize/compress image
  POST /convert-image                   → Convert image format
  POST /crop-image                      → Crop image to dimensions
  POST /compress-files                  → Zip files into archive
  POST /extract-archive                 → Extract zip/archive
  POST /download-url                    → Download file from URL
  POST /find-duplicates                 → Find duplicate files in folder
  POST /batch-rename                    → Rename files matching pattern
  POST /pdf-to-images                   → Render PDF pages as images
  POST /images-to-pdf                   → Combine images into a single PDF
  POST /run-python                      → Execute Python code (Gemini coding power)
  POST /memory                          → Save a personal fact to persistent memory
  GET  /memory/search?q=                → Search memories by keyword
  GET  /memory/list                     → List all memories (optional category filter)
  DELETE /memory/{id}                   → Delete a memory
  GET  /memory/context                  → All memories as text for system prompt
  POST /memory/forget                   → Forget a memory by description (no ID needed)
  POST /transcribe-audio                → Transcribe audio file to text (Gemini)
  POST /search-audio                    → Search within audio file for content (Gemini)
  GET  /search-facts?q=                 → Search extracted facts from documents
  GET  /indexing/status                 → Current indexing progress for active jobs
  GET  /setting?key=                    → Get a setting value
  POST /setting?key=&value=             → Set a setting value
  POST /reminders                        → Save a reminder
  GET  /reminders                        → List reminders
  DELETE /reminders/{id}                 → Delete a reminder
  PUT  /reminders/{id}                   → Reschedule a reminder
  POST /extract-tables                   → Extract structured tables from PDF
  POST /watch?folder=                   → Start auto-indexing a folder
  POST /unwatch?folder=                 → Stop auto-indexing a folder
  GET  /watched                         → List watched folders
  POST /score-photo                     → Score a photo's quality (Gemini vision)
  POST /cull-photos                     → Auto-cull photos in a folder (background)
  GET  /cull-photos/status              → Poll cull job progress
  POST /suggest-categories               → Sample photos + suggest grouping categories (Gemini)
  POST /group-photos                    → Auto-group photos by Gemini vision (background)
  GET  /group-photos/status             → Poll group job progress
"""

import ast
import math
import os
import shutil
import time
from datetime import datetime
from typing import Optional
from dotenv import load_dotenv

load_dotenv()  # Load .env for JINA_API_KEY etc.

from fastapi import FastAPI, Query, HTTPException
from pydantic import BaseModel

import threading

from database import init_db, get_stats, DB_PATH
from search import search
from indexer import index_folder

app = FastAPI(title="Pinpoint", version="0.2.0")

# --- API auth (shared secret) ---
API_SECRET = os.environ.get("API_SECRET", "")

@app.middleware("http")
async def check_api_secret(request, call_next):
    """Require API_SECRET header on all endpoints except /ping."""
    if API_SECRET and request.url.path != "/ping":
        token = request.headers.get("X-API-Secret", "")
        if token != API_SECRET:
            from starlette.responses import JSONResponse
            return JSONResponse({"error": "Unauthorized"}, status_code=401)
    return await call_next(request)

# Thread-local DB connections (one per thread — safe for FastAPI/uvicorn)
_local = threading.local()
_migrations_done = False
_migrations_lock = threading.Lock()


def _get_conn():
    global _migrations_done
    conn = getattr(_local, "conn", None)
    if conn is None:
        conn = init_db(DB_PATH)
        _local.conn = conn
    # Run migrations once across all threads
    if not _migrations_done:
        with _migrations_lock:
            if not _migrations_done:
                try:
                    conn.execute("SELECT superseded_by FROM memories LIMIT 1")
                except Exception:
                    try: conn.execute("ALTER TABLE memories ADD COLUMN superseded_by INTEGER DEFAULT NULL")
                    except Exception: pass
                try:
                    conn.execute("SELECT 1 FROM facts LIMIT 1")
                except Exception:
                    conn.execute("""CREATE TABLE IF NOT EXISTS facts (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        document_id INTEGER NOT NULL,
                        fact_text TEXT NOT NULL,
                        category TEXT DEFAULT 'general',
                        created_at TEXT NOT NULL
                    )""")
                    conn.commit()
                _migrations_done = True
    return conn


def _human_size(size_bytes: int) -> str:
    """Convert bytes to human-readable string."""
    for unit in ("B", "KB", "MB", "GB"):
        if abs(size_bytes) < 1024:
            return f"{size_bytes:.1f} {unit}" if unit != "B" else f"{size_bytes} B"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"


def _human_date(timestamp: float) -> str:
    """Convert Unix timestamp to human-readable date."""
    return datetime.fromtimestamp(timestamp).strftime("%d %b %Y, %I:%M %p")


# --- Blocked paths for safety ---

BLOCKED_PREFIXES = ("/etc", "/usr", "/bin", "/sbin", "/boot", "/proc", "/sys",
                    "/dev", "/var/run", "/var/lock", "C:\\Windows", "C:\\Program Files")


def _is_safe_path(path: str) -> bool:
    """Check if a path is safe to operate on (not a system directory)."""
    abs_path = os.path.abspath(path)
    for prefix in BLOCKED_PREFIXES:
        if abs_path.startswith(prefix):
            return False
    return True


# --- Health check ---

@app.get("/ping")
def ping():
    return {"status": "ok"}


# --- Search (enhanced with filters) ---

@app.get("/search")
def search_endpoint(
    q: str = Query(..., description="Search query"),
    limit: int = Query(20, ge=1, le=100, description="Max results"),
    file_type: Optional[str] = Query(None, description="Filter by type: pdf, docx, xlsx, image, etc."),
    folder: Optional[str] = Query(None, description="Filter by folder path prefix"),
):
    """Search across all indexed documents. Supports file_type and folder filters."""
    result = search(q, DB_PATH, limit, file_type=file_type, folder=folder)
    if not result.get("results"):
        result["_hint"] = "No results. File may not be indexed — use index_file first, then retry. Or try search_facts for quick factual lookups."
    else:
        result["_hint"] = f"{len(result['results'])} result(s) found. Answer the user's question using these."
    return result


# --- Indexing with progress tracking (Segment 18L: Supermemory pattern) ---

_indexing_progress = {}  # folder -> { total, processed, current_file, status, started_at }
_embedding_jobs = {}  # folder → {status, total, done, error?}

class IndexRequest(BaseModel):
    folder: str


@app.post("/index")
def index_endpoint(req: IndexRequest):
    """Index all supported files in a folder. Large folders run in background automatically."""
    import threading
    folder = os.path.abspath(req.folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    # Already running?
    if folder in _indexing_progress and _indexing_progress[folder].get("status") == "indexing":
        p = _indexing_progress[folder]
        return {"status": "indexing", "folder": folder, "total": p["total"], "processed": p["processed"],
                "percent": p["percent"],
                "_hint": f"Already indexing ({p['processed']}/{p['total']}). Tell the user to wait."}

    # Count files to decide sync vs async
    supported_exts = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
                      ".txt", ".csv", ".epub", ".md", ".json", ".log",
                      ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp", ".heic", ".heif"}
    file_count = 0
    for root, dirs, files in os.walk(folder):
        for f in files:
            if os.path.splitext(f)[1].lower() in supported_exts:
                file_count += 1
        if file_count > 50:
            break  # Enough to decide

    if file_count > 50:
        # Large folder — background
        _indexing_progress[folder] = {"total": file_count, "processed": 0, "current_file": "", "status": "indexing", "percent": 0}

        def _bg_index():
            try:
                index_folder(folder, DB_PATH, progress_callback=_update_progress)
                _indexing_progress[folder]["status"] = "done"
                _indexing_progress[folder]["percent"] = 100
            except Exception as e:
                _indexing_progress[folder]["status"] = "error"
                _indexing_progress[folder]["error"] = str(e)

        threading.Thread(target=_bg_index, daemon=True).start()
        return {"status": "indexing", "folder": folder, "total": file_count, "background": True,
                "_hint": f"Indexing {file_count}+ files in background. Tell the user it's started and they can check progress or search once done."}

    # Small folder — sync
    result = index_folder(folder, DB_PATH, progress_callback=_update_progress)
    _indexing_progress.pop(folder, None)
    return result


def _update_progress(folder: str, total: int, processed: int, current_file: str):
    """Callback for indexing progress."""
    _indexing_progress[folder] = {
        "total": total, "processed": processed,
        "current_file": os.path.basename(current_file) if current_file else "",
        "status": "indexing" if processed < total else "done",
        "percent": round(processed / total * 100) if total > 0 else 0,
    }


@app.get("/indexing/status")
def indexing_status():
    """Get current indexing progress for all active jobs. Includes both indexing and embedding."""
    jobs = []
    for k, v in _indexing_progress.items():
        jobs.append({"folder": k, "type": "index", **v})
    for k, v in _embedding_jobs.items():
        jobs.append({"folder": k, "type": "embed", **v})
    if not jobs:
        return {"active": False, "jobs": []}
    return {
        "active": any(j.get("status") in ("indexing", "running") for j in jobs),
        "jobs": jobs,
    }


# --- Stats ---

@app.get("/status")
def status_endpoint():
    """Return indexing statistics."""
    conn = _get_conn()
    stats = get_stats(conn)
    stats["db_path"] = DB_PATH
    stats["db_size_mb"] = round(os.path.getsize(DB_PATH) / 1024 / 1024, 2) if os.path.exists(DB_PATH) else 0
    return stats


# --- Get document by ID ---

@app.get("/document/{doc_id}")
def document_endpoint(doc_id: int):
    """Get full document text by ID. Used by Gemini read_document tool."""
    conn = _get_conn()
    row = conn.execute("""
        SELECT d.id, d.path, d.title, d.file_type, d.page_count, d.active, c.text
        FROM documents d
        JOIN content c ON c.hash = d.hash
        WHERE d.id = ?
    """, (doc_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    return dict(row)


# --- List files (enhanced with sort/filter) ---

# Extension groups for filter_type
_TYPE_EXTENSIONS = {
    "image": {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif", ".tiff", ".tif", ".svg", ".ico", ".heic"},
    "document": {".pdf", ".docx", ".doc", ".txt", ".epub", ".rtf", ".odt"},
    "spreadsheet": {".xlsx", ".xls", ".csv", ".ods"},
    "presentation": {".pptx", ".ppt", ".odp"},
    "video": {".mp4", ".mkv", ".avi", ".mov", ".wmv", ".flv", ".webm"},
    "audio": {".mp3", ".wav", ".flac", ".aac", ".ogg", ".wma", ".m4a"},
    "archive": {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2"},
}


@app.get("/list_files")
def list_files_endpoint(
    folder: str = Query(..., description="Folder path"),
    sort_by: str = Query("name", description="Sort by: name, date, size"),
    filter_ext: Optional[str] = Query(None, description="Filter by extension: .pdf, .xlsx"),
    filter_type: Optional[str] = Query(None, description="Filter by type: image, document, spreadsheet, presentation, video, audio, archive"),
    name_contains: Optional[str] = Query(None, description="Filter by filename containing this text (case-insensitive)"),
    recursive: bool = Query(False, description="Search subdirectories recursively"),
    limit: int = Query(200, ge=1, le=2000, description="Max entries"),
):
    """List contents of a folder with sorting and filtering."""
    folder = os.path.abspath(folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    # Resolve filter_type to a set of extensions
    allowed_exts = None
    if filter_type and filter_type.lower() in _TYPE_EXTENSIONS:
        allowed_exts = _TYPE_EXTENSIONS[filter_type.lower()]
    elif filter_ext:
        allowed_exts = {filter_ext.lower()}

    name_lower = name_contains.lower() if name_contains else None
    entries = []
    _SCAN_CAP = 10000  # cap total files scanned to prevent multi-minute walks

    def _process_entry(de_name, de_path, is_scan_entry=True):
        """Process a directory entry and add to entries if it matches filters."""
        if name_lower and name_lower not in de_name.lower():
            return
        if allowed_exts:
            ext = os.path.splitext(de_name)[1].lower()
            if ext not in allowed_exts:
                return
        try:
            st = os.stat(de_path)
            is_dir = os.path.isdir(de_path)
            if allowed_exts and is_dir:
                return
            entries.append({
                "name": de_name,
                "path": de_path,
                "is_dir": is_dir,
                "size": st.st_size if not is_dir else 0,
                "size_human": _human_size(st.st_size) if not is_dir else "-",
                "modified": _human_date(st.st_mtime),
                "modified_ts": st.st_mtime,
            })
        except OSError:
            pass

    if recursive and name_lower:
        # Use native OS search — fast, no cap, searches everything
        import subprocess, re
        _found_via_native = False
        try:
            # WSL: /mnt/X/ paths → use Windows cmd.exe dir /s /b (NTFS-native, fast)
            # Linux paths → use find command
            wsl_match = re.match(r"^/mnt/([a-zA-Z])/(.*)$", folder)
            if wsl_match:
                drive = wsl_match.group(1).upper()
                win_rest = wsl_match.group(2).rstrip("/").replace("/", "\\")
                win_folder = f"{drive}:\\{win_rest}" if win_rest else f"{drive}:\\"
                dir_cmd = f"dir /s /b {win_folder}\\*{name_contains}*"
                r = subprocess.run(
                    ["/mnt/c/Windows/System32/cmd.exe", "/c", dir_cmd],
                    capture_output=True, text=True, timeout=30
                )
                for line in r.stdout.strip().split("\n"):
                    line = line.strip()
                    if not line:
                        continue
                    # Convert Windows path → WSL: C:\Users\x → /mnt/c/Users/x
                    wsl_path = re.sub(r"^([A-Za-z]):\\", lambda m: f"/mnt/{m.group(1).lower()}/", line)
                    wsl_path = wsl_path.replace("\\", "/")
                    _process_entry(os.path.basename(wsl_path), wsl_path)
                    if len(entries) >= limit:
                        break
                _found_via_native = True
            else:
                cmd = ["find", folder, "-iname", f"*{name_contains}*", "-not", "-path", "*/.*"]
                r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
                for line in r.stdout.strip().split("\n"):
                    if line:
                        _process_entry(os.path.basename(line), line)
                        if len(entries) >= limit:
                            break
                _found_via_native = True
        except Exception:
            pass
        if not _found_via_native:
            # Fallback to os.walk
            scanned = 0
            for root, dirs, files in os.walk(folder):
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                for fname in files:
                    if scanned >= _SCAN_CAP:
                        break
                    _process_entry(fname, os.path.join(root, fname))
                    scanned += 1
                if scanned >= _SCAN_CAP:
                    break
    elif recursive:
        scanned = 0
        for root, dirs, files in os.walk(folder):
            dirs[:] = [d for d in dirs if not d.startswith(".")]
            for fname in files:
                if scanned >= _SCAN_CAP:
                    break
                _process_entry(fname, os.path.join(root, fname))
                scanned += 1
            if scanned >= _SCAN_CAP:
                break
    else:
        # Use os.scandir for performance (caches stat results)
        try:
            with os.scandir(folder) as scanner:
                for de in scanner:
                    _process_entry(de.name, de.path)
        except OSError:
            import subprocess
            try:
                result = subprocess.run(["ls", "-1", folder], capture_output=True, text=True, timeout=10)
                for name in result.stdout.strip().split("\n"):
                    if name:
                        _process_entry(name, os.path.join(folder, name))
            except Exception:
                pass

    # Sort
    if sort_by == "date":
        entries.sort(key=lambda e: e["modified_ts"], reverse=True)
    elif sort_by == "size":
        entries.sort(key=lambda e: e["size"], reverse=True)
    else:
        entries.sort(key=lambda e: (not e["is_dir"], e["name"].lower()))

    total = len(entries)
    entries = entries[:limit]

    result = {"folder": folder, "total": total, "showing": len(entries), "entries": entries}
    # Add summary stats so Gemini can answer without parsing all entries
    if entries and sort_by == "size" and not entries[0].get("is_dir"):
        result["largest"] = f"{entries[0]['name']} ({entries[0]['size_human']})"
    if total > len(entries):
        result["_hint"] = f"{total} items found, showing {len(entries)}. Use name_contains or filter_ext to narrow down."
    elif total > 0:
        result["_hint"] = f"{total} item(s) listed. Use this to answer or plan next action."
    else:
        result["_hint"] = "Folder is empty. Check if the path is correct."
    return result


# --- Grep (content search inside files) ---

class GrepRequest(BaseModel):
    pattern: str
    folder: str
    file_filter: Optional[str] = None  # e.g. "*.txt", "*.xlsx"
    max_results: int = 20

@app.post("/grep")
def grep_endpoint(req: GrepRequest):
    """Search inside files by text pattern using OS grep."""
    import subprocess
    folder = os.path.abspath(req.folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    cmd = ["grep", "-ril", "--include=" + (req.file_filter or "*"), req.pattern, folder]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        files = [f for f in r.stdout.strip().split("\n") if f][:req.max_results]
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=408, detail="Search timed out (folder too large)")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not files:
        return {"pattern": req.pattern, "folder": folder, "matches": 0, "files": []}

    # Get matching lines for context
    results = []
    for fpath in files:
        try:
            r2 = subprocess.run(
                ["grep", "-in", "--max-count=3", req.pattern, fpath],
                capture_output=True, text=True, timeout=5
            )
            lines = [l.strip() for l in r2.stdout.strip().split("\n") if l.strip()][:3]
            results.append({"path": fpath, "name": os.path.basename(fpath), "matches": lines})
        except Exception:
            results.append({"path": fpath, "name": os.path.basename(fpath), "matches": []})

    return {"pattern": req.pattern, "folder": folder, "matches": len(results), "files": results}


# --- File info ---

@app.get("/file_info")
def file_info_endpoint(
    path: str = Query(..., description="File or folder path"),
):
    """Get detailed file/folder metadata and indexed status."""
    path = os.path.abspath(path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Not found: {path}")

    stat = os.stat(path)
    is_dir = os.path.isdir(path)

    info = {
        "name": os.path.basename(path),
        "path": path,
        "is_dir": is_dir,
        "size": stat.st_size,
        "size_human": _human_size(stat.st_size),
        "created": _human_date(stat.st_ctime),
        "modified": _human_date(stat.st_mtime),
        "extension": os.path.splitext(path)[1].lower() if not is_dir else None,
    }

    if is_dir:
        # Count files in directory (single-pass with scandir)
        try:
            file_count = folder_count = 0
            with os.scandir(path) as it:
                for entry in it:
                    if entry.is_file():
                        file_count += 1
                    elif entry.is_dir():
                        folder_count += 1
            info["file_count"] = file_count
            info["folder_count"] = folder_count
        except OSError:
            info["file_count"] = 0
            info["folder_count"] = 0
    else:
        # Check if indexed in database
        conn = _get_conn()
        row = conn.execute(
            "SELECT id, file_type, active FROM documents WHERE path = ?", (path,)
        ).fetchone()
        if row:
            info["indexed"] = True
            info["document_id"] = row["id"]
            info["file_type"] = row["file_type"]
            info["active"] = bool(row["active"])
        else:
            info["indexed"] = False
            info["_hint"] = "File not indexed. Use index_file to make it searchable."

    return info


# --- Calculate (safe math) ---

class CalculateRequest(BaseModel):
    expression: str


# Safe functions/names for math evaluation
_SAFE_MATH = {
    "abs": abs, "round": round, "min": min, "max": max, "sum": sum, "len": len,
    "int": int, "float": float,
    "pi": math.pi, "e": math.e,
    "sqrt": math.sqrt, "ceil": math.ceil, "floor": math.floor,
    "log": math.log, "log10": math.log10, "pow": math.pow,
}


def _safe_eval(expression: str) -> float:
    """Safely evaluate a math expression using AST parsing."""
    # Parse the expression into an AST
    tree = ast.parse(expression, mode="eval")

    # Walk the AST and check all nodes are safe
    for node in ast.walk(tree):
        if isinstance(node, (ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant,
                             ast.Add, ast.Sub, ast.Mult, ast.Div, ast.FloorDiv,
                             ast.Mod, ast.Pow, ast.USub, ast.UAdd,
                             ast.Call, ast.Name, ast.Load, ast.List, ast.Tuple)):
            continue
        raise ValueError(f"Unsupported operation: {type(node).__name__}")

    # Compile and evaluate with only safe builtins
    code = compile(tree, "<calc>", "eval")
    return eval(code, {"__builtins__": {}}, _SAFE_MATH)


@app.post("/calculate")
def calculate_endpoint(req: CalculateRequest):
    """Safely evaluate a mathematical expression."""
    expr = req.expression.strip()
    if not expr:
        raise HTTPException(status_code=400, detail="Empty expression")

    try:
        result = _safe_eval(expr)
        # Format with commas for large numbers
        if isinstance(result, float) and result == int(result) and abs(result) < 1e15:
            formatted = f"{int(result):,}"
        elif isinstance(result, (int,)):
            formatted = f"{result:,}"
        else:
            formatted = f"{result:,.4f}".rstrip("0").rstrip(".")
        return {"expression": expr, "result": result, "formatted": formatted}
    except (ValueError, SyntaxError, TypeError, ZeroDivisionError) as e:
        raise HTTPException(status_code=400, detail=f"Invalid expression: {e}")


# --- Read Excel ---

class ReadExcelRequest(BaseModel):
    path: str
    sheet_name: Optional[str] = None
    cell_range: Optional[str] = None


@app.post("/read_excel")
def read_excel_endpoint(req: ReadExcelRequest):
    """Read specific cells/ranges from an Excel file via openpyxl."""
    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    if not path.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm files supported")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot open Excel file: {e}")

    sheet_names = wb.sheetnames
    result = {"path": path, "sheet_names": sheet_names}

    # Select sheet
    if req.sheet_name:
        if req.sheet_name not in sheet_names:
            wb.close()
            raise HTTPException(status_code=400,
                                detail=f"Sheet '{req.sheet_name}' not found. Available: {sheet_names}")
        ws = wb[req.sheet_name]
    else:
        ws = wb.active

    result["active_sheet"] = ws.title

    if req.cell_range:
        # Read specific range
        try:
            cells = ws[req.cell_range]
        except Exception as e:
            wb.close()
            raise HTTPException(status_code=400, detail=f"Invalid range '{req.cell_range}': {e}")

        # Handle single cell
        if not isinstance(cells, tuple):
            result["data"] = [[str(cells.value) if cells.value is not None else ""]]
            result["rows"] = 1
            result["cols"] = 1
        else:
            # Range of cells — could be rows of tuples or a single tuple
            rows = []
            if isinstance(cells[0], tuple):
                for row in cells:
                    rows.append([str(c.value) if c.value is not None else "" for c in row])
            else:
                # Single row/column
                rows.append([str(c.value) if c.value is not None else "" for c in cells])
            result["data"] = rows
            result["rows"] = len(rows)
            result["cols"] = len(rows[0]) if rows else 0
    else:
        # No range — return first 20 rows
        rows = []
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
            rows.append([str(v) if v is not None else "" for v in row])
        result["data"] = rows
        result["rows"] = len(rows)
        result["cols"] = len(rows[0]) if rows else 0

    # Format as markdown table for readability
    if result.get("data") and len(result["data"]) > 0:
        lines = []
        for i, row in enumerate(result["data"]):
            lines.append("| " + " | ".join(row) + " |")
            if i == 0:
                lines.append("| " + " | ".join(["---"] * len(row)) + " |")
        result["table"] = "\n".join(lines)

    wb.close()
    return result


# --- Move file ---

class MoveFileRequest(BaseModel):
    source: str
    destination: str
    is_copy: bool = False


@app.post("/move_file")
def move_file_endpoint(req: MoveFileRequest):
    """Move, copy, or rename a file."""
    src = os.path.abspath(req.source)
    dest = os.path.abspath(req.destination)

    if not os.path.exists(src):
        raise HTTPException(status_code=404, detail=f"Source not found: {src}")
    if not _is_safe_path(src) or not _is_safe_path(dest):
        raise HTTPException(status_code=403, detail="Cannot operate on system directories")

    # If dest is a directory, move file INTO it
    if os.path.isdir(dest):
        dest = os.path.join(dest, os.path.basename(src))

    # Don't overwrite
    if os.path.exists(dest) and os.path.abspath(src) != os.path.abspath(dest):
        raise HTTPException(status_code=409, detail=f"Destination already exists: {dest}")

    # Create parent directory if needed
    os.makedirs(os.path.dirname(dest), exist_ok=True)

    if req.is_copy:
        if os.path.isdir(src):
            shutil.copytree(src, dest)
        else:
            shutil.copy2(src, dest)
        action = "copied"
    else:
        shutil.move(src, dest)
        action = "moved"

        # Update database path if the file was indexed
        conn = _get_conn()
        conn.execute("UPDATE documents SET path = ? WHERE path = ?", (dest, src))
        conn.commit()

    return {
        "success": True, "source": src, "destination": dest, "action": action,
        "_hint": f"{action.capitalize()} {os.path.basename(src)} → {dest}. Report this to user.",
    }


# --- Batch move/copy files ---

class BatchMoveRequest(BaseModel):
    sources: list[str]
    destination: str
    is_copy: bool = False


@app.post("/batch_move")
def batch_move_endpoint(req: BatchMoveRequest):
    """Move or copy multiple files to a destination folder."""
    dest_folder = os.path.abspath(req.destination)
    os.makedirs(dest_folder, exist_ok=True)

    if not _is_safe_path(dest_folder):
        raise HTTPException(status_code=403, detail="Cannot operate on system directories")

    results = {"moved": [], "skipped": [], "errors": []}
    conn = _get_conn()

    for src_path in req.sources:
        src = os.path.abspath(src_path)
        dest = os.path.join(dest_folder, os.path.basename(src))
        try:
            if not os.path.exists(src):
                results["skipped"].append(os.path.basename(src))
                continue
            if os.path.exists(dest):
                results["skipped"].append(os.path.basename(src))
                continue
            if req.is_copy:
                shutil.copy2(src, dest)
            else:
                shutil.move(src, dest)
                conn.execute("UPDATE documents SET path = ? WHERE path = ?", (dest, src))
            results["moved"].append(os.path.basename(src))
        except Exception as e:
            results["errors"].append(f"{os.path.basename(src)}: {e}")

    conn.commit()
    action = "copied" if req.is_copy else "moved"
    moved_count = len(results["moved"])
    skipped_count = len(results["skipped"])
    error_count = len(results["errors"])
    # Honest result hint — forces truthful reporting (OpenClaw pattern: make lies impossible)
    if moved_count == 0:
        hint = f"WARNING: 0 files were {action}. {skipped_count} skipped (already exist at destination or source missing). {error_count} errors. Tell the user NO files were {action}."
    else:
        hint = f"ACTUALLY {action} {moved_count} files to {dest_folder}. {skipped_count} skipped, {error_count} errors. Report these exact numbers."
    return {
        "success": moved_count > 0,
        "action": action,
        "destination": dest_folder,
        "moved_count": moved_count,
        "skipped_count": skipped_count,
        "error_count": error_count,
        "errors": results["errors"][:5],
        "_hint": hint,
    }


# --- Create folder ---

class CreateFolderRequest(BaseModel):
    path: str


@app.post("/create_folder")
def create_folder_endpoint(req: CreateFolderRequest):
    """Create a new directory (with parents if needed)."""
    path = os.path.abspath(req.path)

    if not _is_safe_path(path):
        raise HTTPException(status_code=403, detail="Cannot create system directories")

    already_existed = os.path.exists(path)
    os.makedirs(path, exist_ok=True)

    return {"success": True, "path": path, "already_existed": already_existed}


# --- Delete file ---

class DeleteFileRequest(BaseModel):
    path: str


@app.post("/delete_file")
def delete_file_endpoint(req: DeleteFileRequest):
    """Delete a file (not folders — too dangerous)."""
    path = os.path.abspath(req.path)

    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Not found: {path}")
    if os.path.isdir(path):
        raise HTTPException(status_code=400, detail="Cannot delete directories (safety). Only files.")
    if not _is_safe_path(path):
        raise HTTPException(status_code=403, detail="Cannot delete system files")

    os.remove(path)

    # Soft-delete from database if indexed
    conn = _get_conn()
    conn.execute("UPDATE documents SET active = 0 WHERE path = ?", (path,))
    conn.commit()

    return {"success": True, "path": path}


# --- Read file from disk (images → base64, docs → text) ---

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif", ".tiff", ".tif", ".heic"}
_TEXT_EXTS = {".txt", ".csv", ".md", ".log", ".json", ".xml", ".yaml", ".yml", ".ini", ".cfg", ".html", ".css", ".js", ".py", ".sh"}
_PDF_EXTS = {".pdf"}
_OFFICE_EXTS = {".docx", ".pptx", ".epub"}
_EXCEL_EXTS = {".xlsx", ".xlsm"}
_MAX_READ_SIZE = 10 * 1024 * 1024  # 10 MB max
_MAX_TEXT_CHARS = 8000  # truncate text for Gemini context


class ReadFileRequest(BaseModel):
    path: str


@app.post("/read_file")
def read_file_endpoint(req: ReadFileRequest):
    """Read actual file from disk. Images return base64, documents return text."""
    import base64

    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    if os.path.isdir(path):
        raise HTTPException(status_code=400, detail="Cannot read a directory")
    if not _is_safe_path(path):
        raise HTTPException(status_code=403, detail="Cannot read system files")

    file_size = os.path.getsize(path)
    if file_size > _MAX_READ_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large ({_human_size(file_size)}). Max {_human_size(_MAX_READ_SIZE)}")

    ext = os.path.splitext(path)[1].lower()

    # Images → base64 for Gemini vision + cached caption if indexed
    if ext in _IMAGE_EXTS:
        with open(path, "rb") as f:
            data = base64.b64encode(f.read()).decode("ascii")
        mime_map = {
            ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
            ".bmp": "image/bmp", ".webp": "image/webp", ".gif": "image/gif",
            ".tiff": "image/tiff", ".tif": "image/tiff", ".heic": "image/heic",
        }
        result = {
            "type": "image",
            "mime_type": mime_map.get(ext, "image/jpeg"),
            "data": data,
            "path": path,
            "size": file_size,
        }
        # DB-first: enrich with cached caption if already indexed (free)
        try:
            conn = _get_conn()
            doc = conn.execute("SELECT hash FROM documents WHERE path = ? AND active = 1", (path,)).fetchone()
            if doc:
                cr = conn.execute("SELECT text FROM content WHERE hash = ?", (doc["hash"],)).fetchone()
                if cr and cr["text"]:
                    result["caption"] = cr["text"]
                    result["_hint"] = "Image has cached caption from index — no need to re-caption."
        except Exception:
            pass
        return result

    # Text files → raw content
    if ext in _TEXT_EXTS:
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read(_MAX_TEXT_CHARS + 100)
            truncated = len(content) > _MAX_TEXT_CHARS
            if truncated:
                content = content[:_MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
            resp = {"type": "text", "content": content, "path": path, "truncated": truncated}
            if truncated:
                resp["_hint"] = "Content truncated. Use index_file + search_documents to find specific sections."
            return resp
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read file: {e}")

    # PDFs → extract with PyMuPDF, fallback to OCR for scanned PDFs
    if ext in _PDF_EXTS:
        try:
            import pymupdf4llm
            content = pymupdf4llm.to_markdown(path)
            # If empty text → likely scanned PDF, try OCR
            if not content or not content.strip():
                ocr_result = _ocr_single(path)
                if ocr_result and ocr_result.get("text"):
                    content = ocr_result["text"]
            truncated = len(content) > _MAX_TEXT_CHARS
            if truncated:
                content = content[:_MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
            resp = {"type": "text", "content": content, "path": path, "truncated": truncated}
            if truncated:
                resp["_hint"] = "Content truncated. Use index_file + search_documents to find specific sections."
            return resp
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read PDF: {e}")

    # Office docs → MarkItDown
    if ext in _OFFICE_EXTS:
        try:
            from markitdown import MarkItDown
            md = MarkItDown(enable_plugins=False)
            result = md.convert(path)
            content = result.text_content
            truncated = len(content) > _MAX_TEXT_CHARS
            if truncated:
                content = content[:_MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
            resp = {"type": "text", "content": content, "path": path, "truncated": truncated}
            if truncated:
                resp["_hint"] = "Content truncated. Use index_file + search_documents to find specific sections."
            return resp
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read document: {e}")

    # Excel → all-sheets summary (sheet names + row counts + first 5 rows each)
    if ext in _EXCEL_EXTS:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_count = ws.max_row or 0
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 6:
                        break
                    rows.append([str(v) if v is not None else "" for v in row])
                lines = [f"## Sheet: {sheet_name} ({row_count} rows)"]
                if rows:
                    lines.append("| " + " | ".join(rows[0]) + " |")
                    lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
                    for row in rows[1:]:
                        padded = row + [""] * (len(rows[0]) - len(row))
                        lines.append("| " + " | ".join(padded[:len(rows[0])]) + " |")
                    if row_count > 6:
                        lines.append(f"... ({row_count - 6} more rows)")
                parts.append("\n".join(lines))
            wb.close()
            content = "\n\n".join(parts)
            return {"type": "text", "content": content, "path": path, "truncated": False,
                    "_hint": "For data analysis (filter, groupby, search values), use analyze_data instead."}
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read Excel: {e}")

    # Unknown type → try as text
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read(_MAX_TEXT_CHARS + 100)
        truncated = len(content) > _MAX_TEXT_CHARS
        if truncated:
            content = content[:_MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
        return {"type": "text", "content": content, "path": path, "truncated": truncated}
    except Exception:
        raise HTTPException(status_code=400, detail=f"Cannot read file type: {ext}")


# --- Conversation memory (Segment 13) ---

class ConversationMessage(BaseModel):
    session_id: str
    role: str
    content: str


class ConversationResetRequest(BaseModel):
    session_id: str


@app.post("/conversation")
def conversation_save(msg: ConversationMessage):
    """Save a conversation message (user or assistant)."""
    if msg.role not in ("user", "assistant"):
        raise HTTPException(status_code=400, detail="role must be 'user' or 'assistant'")
    if not msg.content.strip():
        raise HTTPException(status_code=400, detail="content cannot be empty")

    conn = _get_conn()
    now = datetime.utcnow().isoformat()

    cursor = conn.execute(
        "INSERT INTO conversations(session_id, role, content, timestamp) VALUES (?, ?, ?, ?)",
        (msg.session_id, msg.role, msg.content.strip(), now)
    )

    # Upsert session metadata
    conn.execute("""
        INSERT INTO conversation_sessions(session_id, created_at, updated_at, message_count)
        VALUES (?, ?, ?, 1)
        ON CONFLICT(session_id) DO UPDATE SET
            updated_at = excluded.updated_at,
            message_count = message_count + 1
    """, (msg.session_id, now, now))

    conn.commit()
    return {"success": True, "id": cursor.lastrowid}


@app.get("/conversation/history")
def conversation_history(
    session_id: str = Query(..., description="Session ID (chat JID)"),
    limit: int = Query(20, ge=1, le=100, description="Max messages to return"),
):
    """Load recent conversation history for a session."""
    conn = _get_conn()

    # Get session metadata (for idle timeout check by caller)
    session = conn.execute(
        "SELECT updated_at, message_count FROM conversation_sessions WHERE session_id = ?",
        (session_id,)
    ).fetchone()

    if not session:
        return {"session_id": session_id, "messages": [], "updated_at": None, "message_count": 0}

    # Get last N messages ordered by timestamp
    rows = conn.execute("""
        SELECT role, content, timestamp FROM conversations
        WHERE session_id = ?
        ORDER BY timestamp DESC
        LIMIT ?
    """, (session_id, limit)).fetchall()

    # Reverse so oldest first (chronological order)
    messages = [{"role": r["role"], "content": r["content"], "timestamp": r["timestamp"]} for r in reversed(rows)]

    return {
        "session_id": session_id,
        "messages": messages,
        "updated_at": session["updated_at"],
        "message_count": session["message_count"],
    }


@app.post("/conversation/reset")
def conversation_reset(req: ConversationResetRequest):
    """Delete all messages for a session (reset conversation)."""
    conn = _get_conn()

    cursor = conn.execute(
        "DELETE FROM conversations WHERE session_id = ?", (req.session_id,)
    )
    deleted = cursor.rowcount

    conn.execute(
        "DELETE FROM conversation_sessions WHERE session_id = ?", (req.session_id,)
    )

    conn.commit()
    return {"success": True, "deleted_count": deleted}


@app.get("/conversation/search")
def conversation_search(
    q: str = Query(..., description="Search keywords"),
    session_id: Optional[str] = Query(None, description="Filter by session ID"),
    limit: int = Query(20, ge=1, le=100, description="Max results"),
):
    """Search past conversation messages by keyword."""
    conn = _get_conn()

    if session_id:
        rows = conn.execute("""
            SELECT session_id, role, content, timestamp FROM conversations
            WHERE session_id = ? AND content LIKE ?
            ORDER BY timestamp DESC
            LIMIT ?
        """, (session_id, f"%{q}%", limit)).fetchall()
    else:
        rows = conn.execute("""
            SELECT session_id, role, content, timestamp FROM conversations
            WHERE content LIKE ?
            ORDER BY timestamp DESC
            LIMIT ?
        """, (f"%{q}%", limit)).fetchall()

    results = [dict(r) for r in rows]
    return {"query": q, "count": len(results), "results": results}


# --- Face search (Segment 14) ---

def _get_images_in_folder(folder: str) -> list[str]:
    """List all image files in a folder."""
    folder = os.path.abspath(folder)
    if not os.path.isdir(folder):
        return []
    return sorted([
        os.path.join(folder, f) for f in os.listdir(folder)
        if os.path.splitext(f)[1].lower() in _IMAGE_EXTS
    ])


class DetectFacesRequest(BaseModel):
    image_path: str = None
    folder: str = None


class CropFaceRequest(BaseModel):
    image_path: str
    face_idx: int


class FindPersonRequest(BaseModel):
    reference_image: str
    folder: str
    threshold: float = 0.4


class FindPersonByFaceRequest(BaseModel):
    reference_image: str
    face_idx: int
    folder: str
    threshold: float = 0.4


class CountFacesRequest(BaseModel):
    image_path: str = None
    folder: str = None
    paths: list = None


class CompareFacesRequest(BaseModel):
    image_path_1: str
    face_idx_1: int = 0
    image_path_2: str
    face_idx_2: int = 0


@app.post("/detect-faces")
def detect_faces_endpoint(req: DetectFacesRequest):
    """Detect faces in an image or all images in a folder."""
    from face_search import detect_faces
    conn = _get_conn()

    if req.folder:
        images = _get_images_in_folder(req.folder)
        if not images:
            raise HTTPException(status_code=404, detail=f"No images found in: {req.folder}")
        _BATCH_CAP = 100  # prevent multi-minute blocking on huge folders
        capped = len(images) > _BATCH_CAP
        images = images[:_BATCH_CAP]
        results = {}
        for img_path in images:
            result = detect_faces(img_path, conn)
            if not (isinstance(result, dict) and "error" in result):
                results[os.path.basename(img_path)] = {"faces": result, "face_count": len(result)}
        resp = {"folder": os.path.abspath(req.folder), "images_processed": len(results), "results": results}
        if capped:
            resp["_hint"] = f"Capped at {_BATCH_CAP} images. Process remaining in separate calls."
        else:
            resp["_hint"] = f"Face detection complete for {len(results)} images. Report results directly."
        return resp

    result = detect_faces(req.image_path, conn)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    resp = {"image_path": os.path.abspath(req.image_path), "faces": result, "face_count": len(result)}
    if len(result) > 0:
        resp["_hint"] = "Use find_person(ref_image, folder) to search for this person in other photos. Use crop_face to isolate a specific face."
    return resp


@app.post("/crop-face")
def crop_face_endpoint(req: CropFaceRequest):
    """Crop a specific face from an image. Returns path to cropped image."""
    from face_search import crop_face
    conn = _get_conn()
    result = crop_face(req.image_path, req.face_idx, conn)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@app.post("/find-person")
def find_person_endpoint(req: FindPersonRequest):
    """Find photos matching a reference face in a folder."""
    from face_search import find_person
    conn = _get_conn()
    result = find_person(req.reference_image, req.folder, conn, req.threshold)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    if isinstance(result, list):
        result = {"matches": result, "count": len(result), "_hint": f"{len(result)} photo(s) with matching face. Report or send these."}
    return result


@app.post("/find-person-by-face")
def find_person_by_face_endpoint(req: FindPersonByFaceRequest):
    """Find photos matching a specific face (by index) from a reference image."""
    from face_search import find_person_by_face
    conn = _get_conn()
    result = find_person_by_face(req.reference_image, req.face_idx, req.folder, conn, req.threshold)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    if isinstance(result, list):
        result = {"matches": result, "count": len(result), "_hint": f"{len(result)} photo(s) with matching face. Report or send these."}
    return result


@app.post("/count-faces")
def count_faces_endpoint(req: CountFacesRequest):
    """Count faces in an image, a list of images, or all images in a folder."""
    from face_search import count_faces
    conn = _get_conn()

    # Batch by paths list (e.g. from visual search results)
    if req.paths:
        results = {}
        for img_path in req.paths:
            abs_path = os.path.abspath(img_path)
            if os.path.isfile(abs_path):
                result = count_faces(abs_path, conn)
                if not (isinstance(result, dict) and "error" in result):
                    results[os.path.basename(abs_path)] = result
        return {"images_processed": len(results), "results": results}

    if req.folder:
        images = _get_images_in_folder(req.folder)
        if not images:
            raise HTTPException(status_code=404, detail=f"No images found in: {req.folder}")
        results = {}
        for img_path in images:
            result = count_faces(img_path, conn)
            if not (isinstance(result, dict) and "error" in result):
                results[os.path.basename(img_path)] = result
        resp = {"folder": os.path.abspath(req.folder), "images_processed": len(results), "results": results}
        resp["_hint"] = f"Face counts complete for {len(results)} images. Report results directly."
        return resp

    result = count_faces(req.image_path, conn)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    result["_hint"] = "Face count complete. Report the number directly."
    return result


@app.post("/compare-faces")
def compare_faces_endpoint(req: CompareFacesRequest):
    """Compare two faces to check if they're the same person."""
    from face_search import compare_faces
    conn = _get_conn()
    result = compare_faces(req.image_path_1, req.face_idx_1,
                           req.image_path_2, req.face_idx_2, conn)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    result["_hint"] = "Score >0.6 = likely same person, 0.4-0.6 = uncertain, <0.4 = different person."
    return result


# --- Face memory (Segment 18V: persistent face recognition) ---

class RememberFaceRequest(BaseModel):
    image_path: str
    face_idx: int = 0
    name: str


class ForgetFaceRequest(BaseModel):
    name: str


class RecognizeFacesRequest(BaseModel):
    image_path: str


@app.post("/remember-face")
def remember_face_endpoint(req: RememberFaceRequest):
    """Save a face for future recognition."""
    from face_search import remember_face
    conn = _get_conn()
    result = remember_face(req.image_path, req.face_idx, req.name, conn)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    result["_hint"] = f"Face saved as '{req.name}'. detect_faces will now auto-recognize this person. Add more photos of them for better accuracy."
    return result


@app.post("/forget-face")
def forget_face_endpoint(req: ForgetFaceRequest):
    """Delete all saved face data for a person."""
    from face_search import forget_face
    conn = _get_conn()
    result = forget_face(req.name, conn)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@app.post("/recognize-faces")
def recognize_faces_endpoint(req: RecognizeFacesRequest):
    """Recognize known faces in an image."""
    from face_search import recognize_faces
    conn = _get_conn()
    result = recognize_faces(req.image_path, conn)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


# --- Visual image search (Segment 18C: SigLIP2 embeddings) ---

class VisualSearchRequest(BaseModel):
    folder: str
    query: str
    limit: int = 10
    recursive: bool = False


@app.post("/search-images-visual")
def search_images_visual_endpoint(req: VisualSearchRequest):
    """Search images in a folder by text description using SigLIP2 embeddings."""
    folder = os.path.abspath(req.folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=404, detail=f"Folder not found: {folder}")

    from image_search import search_images, _get_image_files, _mem_cache, _load_cached_embeddings, _HAS_SIGLIP

    # Check if embedding is needed (folder not in memory cache)
    # Auto-recurse if no images at top level but subfolders exist
    recursive = req.recursive
    files = _get_image_files(folder, recursive=recursive)
    if not files and not recursive:
        files = _get_image_files(folder, recursive=True)
        if files:
            recursive = True
    if not files:
        return {"error": f"No images found in {folder}", "results": [],
                "_hint": "No images found. Try search_documents(query=..., file_type='image', folder=...) to search indexed image captions instead."}
    # If auto-recursed into many images, suggest search_documents first (free + instant)
    if recursive and len(files) > 200:
        return {"error": f"Folder has {len(files)} images across subfolders — too many for visual search. Use search_documents(query=..., file_type='image', folder='{folder}') instead — it searches indexed image captions instantly for free.",
                "results": [], "total_images": len(files),
                "_hint": f"Too many images ({len(files)}). Try search_documents with file_type='image' first — it's free and instant."}

    # No SigLIP — dispatch to Gemini vision (no embedding needed)
    if not _HAS_SIGLIP:
        from image_search import _search_images_gemini
        if len(files) > 500:
            # Large folder — run in background
            if folder in _embedding_jobs:
                job = _embedding_jobs[folder]
                if job["status"] == "running":
                    return {"status": "scoring", "total_batches": job["total"], "done_batches": job["done"],
                            "_hint": f"Still scoring images with Gemini ({job['done']}/{job['total']} batches). Tell the user to wait and try again."}
                if job["status"] == "done":
                    result = job.get("result", {})
                    del _embedding_jobs[folder]
                    result["_hint"] = "Visual search complete. Results are AI-analyzed — trust them to answer, categorize, or group."
                    return result
                if job["status"] == "error":
                    del _embedding_jobs[folder]

            import threading, math
            total_batches = math.ceil(len(files) / 200)
            _embedding_jobs[folder] = {"status": "running", "total": total_batches, "done": 0}

            def _bg_gemini_search():
                try:
                    def _progress(done, total):
                        _embedding_jobs[folder]["done"] = done
                    result = _search_images_gemini(folder, req.query, limit=req.limit, progress_callback=_progress, recursive=recursive)
                    _embedding_jobs[folder]["status"] = "done"
                    _embedding_jobs[folder]["result"] = result
                except Exception as e:
                    _embedding_jobs[folder]["status"] = "error"
                    _embedding_jobs[folder]["error"] = str(e)

            threading.Thread(target=_bg_gemini_search, daemon=True).start()
            return {"status": "scoring", "total_images": len(files), "total_batches": total_batches,
                    "_hint": f"Scoring {len(files)} images with Gemini vision in {total_batches} batches. Tell the user it's processing and will be ready soon."}

        # Small folder — do inline
        result = search_images(folder, req.query, limit=req.limit, recursive=recursive)
        if "error" in result and not result.get("results"):
            raise HTTPException(status_code=404, detail=result["error"])
        result["_hint"] = "Visual search complete. Results are AI-analyzed — trust them to answer, categorize, or group."
        return result

    mem = _mem_cache.get(folder)
    if mem and len(mem["paths"]) == len(files):
        # Already cached — search instantly
        result = search_images(folder, req.query, limit=req.limit, recursive=recursive)
        if "error" in result and not result.get("results"):
            raise HTTPException(status_code=404, detail=result["error"])
        return result

    # Check how many need embedding
    cached = _load_cached_embeddings(files)
    to_embed = len(files) - len(cached)

    if to_embed > 50:
        # Large job — run in background
        if folder in _embedding_jobs:
            job = _embedding_jobs[folder]
            if job["status"] == "running":
                return {"status": "embedding", "total": job["total"], "done": job["done"],
                        "_hint": f"Still embedding ({job['done']}/{job['total']}). Tell the user to wait and try again in a bit."}
            if job["status"] == "done":
                # Background embedding finished — search now
                result = search_images(folder, req.query, limit=req.limit, recursive=recursive)
                if "error" in result and not result.get("results"):
                    raise HTTPException(status_code=404, detail=result["error"])
                result["_hint"] = "Visual search complete. Results are AI-analyzed — trust them to answer, categorize, or group."
                return result
            if job["status"] == "error":
                del _embedding_jobs[folder]  # Clear failed job, will retry below

        import threading
        _embedding_jobs[folder] = {"status": "running", "total": len(files), "done": len(cached)}

        def _bg_embed():
            try:
                from image_search import embed_images
                def _progress(done, total):
                    _embedding_jobs[folder]["done"] = done + len(cached)
                embed_images(folder, progress_callback=_progress)
                _embedding_jobs[folder]["status"] = "done"
                _embedding_jobs[folder]["done"] = len(files)
            except Exception as e:
                _embedding_jobs[folder]["status"] = "error"
                _embedding_jobs[folder]["error"] = str(e)

        threading.Thread(target=_bg_embed, daemon=True).start()
        return {"status": "embedding", "total": len(files), "to_embed": to_embed, "cached": len(cached),
                "_hint": f"Embedding {to_embed} images in background. Tell the user it's processing and will be ready in a few minutes. They can search once done."}

    # Small job — do inline
    result = search_images(folder, req.query, limit=req.limit, recursive=recursive)
    if "error" in result and not result.get("results"):
        raise HTTPException(status_code=404, detail=result["error"])
    result["_hint"] = "Visual search complete. Results are AI-analyzed — trust them to answer, categorize, or group."
    return result


@app.get("/embedding-status")
def embedding_status(folder: str = Query(...)):
    """Check background embedding progress."""
    folder = os.path.abspath(folder)
    job = _embedding_jobs.get(folder)
    if not job:
        return {"status": "none", "folder": folder}
    return {"folder": folder, **job}


# --- Video search (Segment 18H: SigLIP2 + FFmpeg) ---

class VideoSearchRequest(BaseModel):
    video_path: str
    query: str
    fps: float = 1.0
    limit: int = 5


@app.post("/search-video")
def search_video_endpoint(req: VideoSearchRequest):
    """Search inside a video by text description using SigLIP2 frame embeddings."""
    video_path = os.path.abspath(req.video_path)
    if not os.path.isfile(video_path):
        raise HTTPException(status_code=404, detail=f"Video not found: {video_path}")

    from video_search import search_video
    result = search_video(video_path, req.query, fps=req.fps, limit=req.limit)
    if "error" in result and not result.get("results"):
        raise HTTPException(status_code=400, detail=result["error"])
    return result


class ExtractFrameRequest(BaseModel):
    video_path: str
    seconds: float
    output_path: str = None


@app.post("/extract-frame")
def extract_frame_endpoint(req: ExtractFrameRequest):
    """Extract a single frame from a video at a given timestamp."""
    video_path = os.path.abspath(req.video_path)
    if not os.path.isfile(video_path):
        raise HTTPException(status_code=404, detail=f"Video not found: {video_path}")

    from video_search import extract_frame_image
    try:
        out = extract_frame_image(video_path, req.seconds, req.output_path)
        return {"path": out}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Frame extraction failed: {e}")


# --- Audio (Gemini native) ---

class TranscribeAudioRequest(BaseModel):
    path: str


@app.post("/transcribe-audio")
def transcribe_audio_endpoint(req: TranscribeAudioRequest):
    """Transcribe an audio file to text using Gemini."""
    path = os.path.abspath(req.path)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail=f"Audio file not found: {path}")
    from audio_search import transcribe_audio
    result = transcribe_audio(path)
    if "error" in result and "text" not in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


class SearchAudioRequest(BaseModel):
    audio_path: str
    query: str
    limit: int = 5


@app.post("/search-audio")
def search_audio_endpoint(req: SearchAudioRequest):
    """Search within an audio file for specific content."""
    path = os.path.abspath(req.audio_path)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail=f"Audio file not found: {path}")
    from audio_search import search_audio
    result = search_audio(path, req.query, limit=req.limit)
    if "error" in result and not result.get("results"):
        raise HTTPException(status_code=400, detail=result["error"])
    return result


# --- OCR (Tesseract) ---

class OcrRequest(BaseModel):
    path: str = None
    folder: str = None


def _ocr_single(path: str) -> dict:
    """OCR a single file (image or PDF). Checks DB first — if already indexed, returns cached text."""
    path = os.path.abspath(path)
    if not os.path.exists(path):
        return {"error": f"File not found: {path}"}

    # DB-first: check if already indexed (free — skip expensive OCR)
    try:
        conn = _get_conn()
        doc = conn.execute("SELECT hash FROM documents WHERE path = ? AND active = 1", (path,)).fetchone()
        if doc:
            content_row = conn.execute("SELECT text FROM content WHERE hash = ?", (doc["hash"],)).fetchone()
            if content_row and content_row["text"] and len(content_row["text"].strip()) > 10:
                return {"path": path, "text": content_row["text"], "method": "cached_index",
                        "_hint": "Text from DB index (no re-processing needed)."}
    except Exception:
        pass

    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        from extractors import extract_pdf
        result = extract_pdf(path)
        if result is None:
            return {"error": "Failed to extract text from PDF"}
        return {"path": path, "text": result["text"], "pages": result.get("page_count", 0), "method": "pdf_ocr"}
    elif ext in _IMAGE_EXTS:
        try:
            from extractors import _ocr_tesseract, _ocr_gemini, _HAS_TESSERACT, _preprocess_image, _get_gemini
            from PIL import Image
            img = Image.open(path).convert("RGB")
            img = _preprocess_image(img)
            _gemini = _get_gemini()
            text = _ocr_gemini([img]) if _gemini else (_ocr_tesseract([img]) if _HAS_TESSERACT else "")
            method = "gemini_ocr" if _gemini else ("tesseract_ocr" if _HAS_TESSERACT else "none")
            return {"path": path, "text": text, "method": method,
                    "_hint": "Use index_file to make this text searchable, or search_documents if already indexed."}
        except Exception as e:
            return {"error": f"OCR failed: {e}"}
    else:
        return {"error": f"OCR not supported for {ext}. Use images or PDFs."}


@app.post("/ocr")
def ocr_endpoint(req: OcrRequest):
    """Extract text from an image, PDF, or all files in a folder using OCR."""
    if req.folder:
        folder = os.path.abspath(req.folder)
        if not os.path.isdir(folder):
            raise HTTPException(status_code=404, detail=f"Folder not found: {folder}")
        ocr_exts = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif", ".pdf"}
        files = sorted([
            os.path.join(folder, f) for f in os.listdir(folder)
            if os.path.splitext(f)[1].lower() in ocr_exts
        ])
        if not files:
            raise HTTPException(status_code=404, detail=f"No images or PDFs found in: {folder}")
        _BATCH_CAP = 50  # prevent multi-minute blocking on huge folders
        capped = len(files) > _BATCH_CAP
        files = files[:_BATCH_CAP]
        results = {}
        for fpath in files:
            result = _ocr_single(fpath)
            if "error" not in result:
                results[os.path.basename(fpath)] = result.get("text", "")
        resp = {"folder": folder, "files_processed": len(results), "results": results}
        if capped:
            resp["_hint"] = f"Capped at {_BATCH_CAP} files. Process remaining files in separate calls."
        return resp

    if not req.path:
        raise HTTPException(status_code=400, detail="Provide path or folder")
    result = _ocr_single(req.path)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


# --- Pandas data analysis (Segment 15 + 18Q: smart cache, multi-sheet, search) ---

# DataFrame LRU cache: (path, mtime, sheet) → DataFrame
_df_cache = {}  # key → {"df": DataFrame, "atime": float}
_DF_CACHE_MAX = 5


def _normalize_for_search(val: str) -> str:
    """Normalize a value for fuzzy matching: strip non-digits if numeric-ish, lowercase."""
    import re
    stripped = re.sub(r"[^\d]", "", val)
    if len(stripped) >= 6:  # likely a phone/ID number
        return stripped
    return val.lower().strip()


def _load_df(path: str, ext: str, sheet: str = None):
    """Load DataFrame with LRU cache. Returns (df, sheet_name, all_sheet_names)."""
    import pandas as pd

    mtime = os.path.getmtime(path)
    cache_key = f"{path}:{mtime}:{sheet or '_default_'}"

    # Check cache
    if cache_key in _df_cache:
        entry = _df_cache[cache_key]
        entry["atime"] = time.time()
        return entry["df"], entry["sheet_name"], entry["all_sheets"]

    # Load fresh
    all_sheets = None
    sheet_name = sheet

    if ext == ".csv":
        df = pd.read_csv(path)
        all_sheets = ["Sheet1"]
        sheet_name = "Sheet1"
    elif ext in (".xlsx", ".xlsm", ".xls"):
        # Get all sheet names first
        xls = pd.ExcelFile(path)
        all_sheets = xls.sheet_names
        if sheet and sheet in all_sheets:
            df = pd.read_excel(xls, sheet_name=sheet)
            sheet_name = sheet
        else:
            df = pd.read_excel(xls, sheet_name=0)
            sheet_name = all_sheets[0]
        xls.close()
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {ext}. Use CSV or Excel.")

    # Evict oldest if cache full
    if len(_df_cache) >= _DF_CACHE_MAX:
        oldest_key = min(_df_cache, key=lambda k: _df_cache[k]["atime"])
        del _df_cache[oldest_key]

    _df_cache[cache_key] = {"df": df, "sheet_name": sheet_name, "all_sheets": all_sheets, "atime": time.time()}
    return df, sheet_name, all_sheets


def _get_schema(df) -> dict:
    """Return column types + sample values for Gemini guidance."""
    schema = {}
    for col in df.columns:
        dtype = str(df[col].dtype)
        sample = None
        non_null = df[col].dropna()
        if len(non_null) > 0:
            sample = str(non_null.iloc[0])
            if len(sample) > 50:
                sample = sample[:50] + "..."
        schema[col] = {"type": dtype, "sample": sample}
    return schema


class AnalyzeDataRequest(BaseModel):
    path: str
    operation: str = "describe"
    columns: Optional[str] = None
    query: Optional[str] = None
    sheet: Optional[str] = None
    head: int = 20


@app.post("/analyze-data")
def analyze_data_endpoint(req: AnalyzeDataRequest):
    """Run pandas analysis on CSV or Excel files. Cached, multi-sheet, with search."""
    import pandas as pd

    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    ext = os.path.splitext(path)[1].lower()
    try:
        df, sheet_name, all_sheets = _load_df(path, ext, req.sheet)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read file: {e}")

    result = {
        "path": path, "shape": list(df.shape), "columns": list(df.columns),
        "sheet": sheet_name, "all_sheets": all_sheets,
    }

    op = req.operation.lower()

    if op == "describe":
        result["data"] = df.describe(include="all").to_string()
        result["schema"] = _get_schema(df)
    elif op == "head":
        result["data"] = df.head(req.head).to_string()
    elif op == "columns":
        result["data"] = _get_schema(df)
    elif op == "search":
        # Vectorized search across cells with normalization (fast even on 100MB files)
        import re as _re
        if not req.query:
            raise HTTPException(status_code=400, detail="query required for search")
        needle = _normalize_for_search(req.query)
        matches = []

        # Determine sheets to search
        sheets_to_search = []
        if ext in (".xlsx", ".xlsm", ".xls") and not req.sheet:
            # Search ALL sheets — load each via cache
            for sn in all_sheets:
                try:
                    sdf, _, _ = _load_df(path, ext, sn)
                    sheets_to_search.append((sn, sdf))
                except Exception:
                    continue
        else:
            sheets_to_search = [(sheet_name, df)]

        for sn, sdf in sheets_to_search:
            if len(matches) >= req.head:
                break
            # Reset index for positional access (avoids slow get_loc on non-default indices)
            sdf_reset = sdf.reset_index(drop=True)
            for col in sdf_reset.columns:
                if len(matches) >= req.head:
                    break
                col_str = sdf_reset[col].astype(str).fillna("")
                if needle.isdigit() and len(needle) >= 6:
                    col_normalized = col_str.str.replace(r"[^\d]", "", regex=True)
                else:
                    col_normalized = col_str.str.lower()
                hit_mask = col_normalized.str.contains(needle, na=False, regex=False)
                hit_positions = hit_mask[hit_mask].index.tolist()  # Already positional after reset_index
                for row_pos in hit_positions:
                    if len(matches) >= req.head:
                        break
                    start = max(0, row_pos - 2)
                    end = min(len(sdf_reset), row_pos + 3)
                    # Build context as dict list (cheaper than to_string per hit)
                    context = sdf_reset.iloc[start:end].to_dict(orient="records")
                    matches.append({
                        "sheet": sn, "row": row_pos + 1, "column": col,
                        "value": str(col_str.iloc[row_pos]), "context": context,
                    })

        result["data"] = matches
        result["matched"] = len(matches)
        if not matches:
            result["_hint"] = f"No matches for '{req.query}' across {len(sheets_to_search)} sheet(s). Try different keywords or check the sheet names."
    elif op == "value_counts":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for value_counts")
        col = req.columns.strip()
        if col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{col}' not found. Available: {list(df.columns)}")
        result["data"] = df[col].value_counts().head(req.head).to_string()
    elif op == "groupby":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for groupby (format: 'group_col:agg_col')")
        parts = req.columns.split(":")
        group_col = parts[0].strip()
        agg_col = parts[1].strip() if len(parts) > 1 else None
        if group_col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{group_col}' not found")
        if agg_col and agg_col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{agg_col}' not found")
        if agg_col:
            result["data"] = df.groupby(group_col)[agg_col].agg(["count", "sum", "mean", "min", "max"]).to_string()
        else:
            result["data"] = df.groupby(group_col).size().to_string()
    elif op == "filter":
        if not req.query:
            raise HTTPException(status_code=400, detail="query required for filter (e.g. 'amount > 1000')")
        try:
            filtered = df.query(req.query)
            n_matched = len(filtered)
            result["data"] = filtered.head(req.head).to_string()
            result["matched_rows"] = n_matched
            # Auto-context: show surrounding rows for sparse matches
            if 0 < n_matched <= 5:
                context_indices = set()
                for idx in filtered.index[:5]:
                    pos = df.index.get_loc(idx)
                    for offset in range(-2, 3):
                        if 0 <= pos + offset < len(df):
                            context_indices.add(df.index[pos + offset])
                result["context"] = df.loc[sorted(context_indices)].to_string()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid query: {e}")
    elif op == "corr":
        numeric_df = df.select_dtypes(include="number")
        result["data"] = numeric_df.corr().to_string()
    elif op == "sort":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for sort")
        col = req.columns.strip()
        ascending = not col.startswith("-")
        col = col.lstrip("-")
        if col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{col}' not found")
        result["data"] = df.sort_values(col, ascending=ascending).head(req.head).to_string()
    elif op == "unique":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for unique")
        col = req.columns.strip()
        if col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{col}' not found")
        uniq = df[col].dropna().unique().tolist()
        result["data"] = uniq[:100]
        result["total_unique"] = len(uniq)
    elif op == "eval":
        if not req.query:
            raise HTTPException(status_code=400, detail="query required for eval (e.g. '(Qty * Price).sum()')")
        try:
            eval_result = eval(f"df.{req.query}" if not req.query.strip().startswith("df") else req.query,
                               {"__builtins__": {}}, {"df": df, "pd": pd})
            if hasattr(eval_result, 'to_string'):
                result["data"] = eval_result.head(req.head).to_string() if hasattr(eval_result, 'head') else eval_result.to_string()
            else:
                result["data"] = str(eval_result)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Eval error: {e}")
    elif op == "shape":
        result["data"] = f"{df.shape[0]} rows x {df.shape[1]} columns"
    else:
        raise HTTPException(status_code=400, detail=f"Unknown operation: {op}. Use: describe, head, columns, value_counts, groupby, filter, corr, sort, unique, eval, search, shape")

    # Sufficiency hints per operation
    if op == "columns":
        result["_hint"] = "Column info loaded. Now call the specific operation you need (filter, groupby, search, sort, etc.)."
    elif op in ("filter", "search", "groupby", "sort", "value_counts", "head", "describe", "eval", "unique", "corr"):
        if not result.get("_hint"):
            result["_hint"] = "Data retrieved. Answer the user's question with these results."

    return result


# --- Index file on demand (Segment 15) ---

class IndexFileRequest(BaseModel):
    path: str


@app.post("/index-file")
def index_file_endpoint(req: IndexFileRequest):
    """Index a single file into the search database on demand. Extracts facts if text is substantial."""
    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    if os.path.isdir(path):
        raise HTTPException(status_code=400, detail="Use POST /index for folders. This endpoint is for single files.")

    from extractors import extract
    from database import upsert_document, chunk_document, content_hash

    # DB-first: skip extraction if file unchanged (same mtime = same content)
    conn = _get_conn()
    existing = conn.execute(
        "SELECT hash, modified_at FROM documents WHERE path = ? AND active = 1", (path,)
    ).fetchone()
    if existing:
        file_mtime = datetime.fromtimestamp(os.path.getmtime(path)).isoformat()
        if existing["modified_at"] and existing["modified_at"] >= file_mtime:
            return {
                "success": True, "path": path, "already_indexed": True,
                "hash": existing["hash"][:16],
                "_hint": "File already indexed (unchanged). Use search_documents to search it.",
            }

    result = extract(path)
    if result is None:
        raise HTTPException(status_code=400, detail=f"Cannot extract text from {os.path.basename(path)}")

    conn = _get_conn()
    h = upsert_document(conn, path, result["text"], result["file_type"], result.get("page_count", 0))

    # Chunk the document for section-level search (Segment 18P)
    chunks_count = 0
    doc_row = conn.execute("SELECT id FROM documents WHERE path = ?", (path,)).fetchone()
    if doc_row:
        doc_id = doc_row["id"]
        try:
            chunks_count = chunk_document(conn, doc_id, result["text"])
        except Exception as e:
            print(f"[Chunk] Failed for {path}: {e}")

    # Extract facts from substantial documents (Supermemory pattern)
    facts_count = 0
    text = result["text"]
    if len(text) > 200:  # Only for non-trivial documents
        if doc_row:
            # Check if facts already extracted for this document
            existing = conn.execute("SELECT COUNT(*) as c FROM facts WHERE document_id = ?", (doc_id,)).fetchone()
            if existing["c"] == 0:
                try:
                    facts_count = _extract_facts(conn, doc_id, text[:4000], os.path.basename(path))
                except Exception as e:
                    print(f"[Facts] Extraction failed for {path}: {e}")

    return {
        "success": True,
        "path": path,
        "file_type": result["file_type"],
        "text_length": len(text),
        "hash": h[:16],
        "chunks": chunks_count,
        "facts_extracted": facts_count,
        "_hint": "File is now searchable. Use search_documents to find specific sections within it.",
    }


def _extract_facts(conn, doc_id: int, text: str, filename: str) -> int:
    """Extract key facts from document text using Gemini."""
    GEMINI_KEY = os.getenv("GEMINI_API_KEY")
    if not GEMINI_KEY:
        return 0

    prompt = f"""Extract 3-10 key facts from this document. One fact per line, no numbering.
Include: names, dates, amounts, topics, key details.
Keep each fact short (1 sentence max).
File: {filename}

Text:
{text}

Facts:"""

    try:
        from google.genai import types as genai_types
        from extractors import _get_gemini, gemini_call_with_retry
        _facts_schema = {
            "type": "OBJECT",
            "properties": {"facts": {"type": "ARRAY", "items": {"type": "STRING"}}},
            "required": ["facts"],
        }
        client = _get_gemini()
        if not client:
            return 0
        response = gemini_call_with_retry(
            client,
            model=os.environ.get("GEMINI_MODEL", "gemini-3.1-flash-lite-preview"),
            contents=prompt,
            config=genai_types.GenerateContentConfig(
                response_mime_type="application/json",
                response_json_schema=_facts_schema,
            ),
        )
        import json as _json_mod
        data = _json_mod.loads(response.text)
        facts_list = data.get("facts", [])
        if not facts_list:
            return 0

        now = datetime.utcnow().isoformat()
        count = 0
        for fact in facts_list:
            fact = str(fact).strip()
            if fact and len(fact) > 10:
                conn.execute(
                    "INSERT INTO facts(document_id, fact_text, created_at) VALUES (?, ?, ?)",
                    (doc_id, fact, now)
                )
                count += 1
        conn.commit()
        return count
    except Exception as e:
        print(f"[Facts] Gemini error: {e}")
        return 0


@app.get("/search-facts")
def search_facts_endpoint(
    q: str = Query(..., description="Search query"),
    limit: int = Query(10, ge=1, le=50),
):
    """Search extracted facts from documents."""
    conn = _get_conn()
    rows = conn.execute(
        """SELECT f.id, f.fact_text, f.category, d.path, d.file_type
           FROM facts f JOIN documents d ON f.document_id = d.id
           WHERE f.fact_text LIKE ? LIMIT ?""",
        (f"%{q}%", limit)
    ).fetchall()
    resp = {"query": q, "count": len(rows), "results": [dict(r) for r in rows]}
    if not rows:
        resp["_hint"] = "No facts match. Try search_documents for full-text search across document content."
    else:
        resp["_hint"] = f"{len(rows)} fact(s) found. Answer the user's question using these."
    return resp


# --- Write & Create (Segment 15B) ---

class WriteFileRequest(BaseModel):
    path: str
    content: str
    append: bool = False


@app.post("/write-file")
def write_file_endpoint(req: WriteFileRequest):
    """Create or write a text file."""
    path = os.path.abspath(req.path)
    if not _is_safe_path(path):
        raise HTTPException(status_code=403, detail="Cannot write to system directories")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    mode = "a" if req.append else "w"
    with open(path, mode, encoding="utf-8") as f:
        f.write(req.content)
    return {"success": True, "path": path, "size": os.path.getsize(path), "append": req.append}


class GenerateExcelRequest(BaseModel):
    path: str
    data: list  # list of dicts or list of lists
    columns: Optional[list] = None
    sheet_name: str = "Sheet1"


@app.post("/generate-excel")
def generate_excel_endpoint(req: GenerateExcelRequest):
    """Create an Excel file from data."""
    import pandas as pd
    path = os.path.abspath(req.path)
    if not _is_safe_path(path):
        raise HTTPException(status_code=403, detail="Cannot write to system directories")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    try:
        df = pd.DataFrame(req.data, columns=req.columns)
        df.to_excel(path, sheet_name=req.sheet_name, index=False)
        return {"success": True, "path": path, "rows": len(df), "columns": list(df.columns)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to create Excel: {e}")


class GenerateChartRequest(BaseModel):
    data: dict  # {"labels": [...], "values": [...]} or {"x": [...], "y": [...]}
    chart_type: str = "bar"  # bar, line, pie, scatter, hist
    title: str = ""
    xlabel: str = ""
    ylabel: str = ""
    output_path: Optional[str] = None


@app.post("/generate-chart")
def generate_chart_endpoint(req: GenerateChartRequest):
    """Generate a chart image using matplotlib."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import tempfile

    output = os.path.abspath(req.output_path) if req.output_path else os.path.join(
        tempfile.gettempdir(), "pinpoint_charts", f"chart_{int(__import__('time').time())}.png"
    )
    os.makedirs(os.path.dirname(output), exist_ok=True)

    try:
        fig, ax = plt.subplots(figsize=(10, 6))
        labels = req.data.get("labels") or req.data.get("x", [])
        values = req.data.get("values") or req.data.get("y", [])

        if req.chart_type == "bar":
            ax.bar(labels, values)
        elif req.chart_type == "line":
            ax.plot(labels, values, marker="o")
        elif req.chart_type == "pie":
            ax.pie(values, labels=labels, autopct="%1.1f%%")
        elif req.chart_type == "scatter":
            ax.scatter(labels, values)
        elif req.chart_type == "hist":
            ax.hist(values, bins="auto")
        else:
            raise HTTPException(status_code=400, detail=f"Unknown chart type: {req.chart_type}. Use: bar, line, pie, scatter, hist")

        if req.title:
            ax.set_title(req.title)
        if req.xlabel:
            ax.set_xlabel(req.xlabel)
        if req.ylabel:
            ax.set_ylabel(req.ylabel)
        plt.tight_layout()
        fig.savefig(output, dpi=150)
        plt.close(fig)
        return {"success": True, "path": output}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Chart generation failed: {e}")


# --- PDF Tools (Segment 15B) ---

class MergePdfRequest(BaseModel):
    paths: list[str]
    output_path: str


@app.post("/merge-pdf")
def merge_pdf_endpoint(req: MergePdfRequest):
    """Merge multiple PDFs into one."""
    import fitz  # PyMuPDF
    output = os.path.abspath(req.output_path)
    if not _is_safe_path(output):
        raise HTTPException(status_code=403, detail="Cannot write to system directories")
    os.makedirs(os.path.dirname(output), exist_ok=True)

    merged = fitz.open()
    for p in req.paths:
        p = os.path.abspath(p)
        if not os.path.exists(p):
            raise HTTPException(status_code=404, detail=f"File not found: {p}")
        doc = fitz.open(p)
        merged.insert_pdf(doc)
        doc.close()
    merged.save(output)
    total_pages = len(merged)
    merged.close()
    return {"success": True, "path": output, "total_pages": total_pages, "files_merged": len(req.paths)}


class SplitPdfRequest(BaseModel):
    path: str
    pages: str  # "1-5", "3,7,10", "1-3,5,8-10"
    output_path: str


@app.post("/split-pdf")
def split_pdf_endpoint(req: SplitPdfRequest):
    """Extract specific pages from a PDF."""
    import fitz
    path = os.path.abspath(req.path)
    output = os.path.abspath(req.output_path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    if not _is_safe_path(output):
        raise HTTPException(status_code=403, detail="Cannot write to system directories")
    os.makedirs(os.path.dirname(output), exist_ok=True)

    # Parse page specification
    page_nums = set()
    for part in req.pages.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            for i in range(int(start), int(end) + 1):
                page_nums.add(i)
        else:
            page_nums.add(int(part))

    doc = fitz.open(path)
    # Convert 1-based to 0-based
    zero_based = sorted(p - 1 for p in page_nums if 1 <= p <= len(doc))
    if not zero_based:
        doc.close()
        raise HTTPException(status_code=400, detail=f"No valid pages. PDF has {len(doc)} pages.")

    new_doc = fitz.open()
    for page_no in zero_based:
        new_doc.insert_pdf(doc, from_page=page_no, to_page=page_no)
    new_doc.save(output)
    extracted = len(new_doc)
    source_pages = len(doc)
    new_doc.close()
    doc.close()
    return {"success": True, "path": output, "pages_extracted": extracted, "source_pages": source_pages}


class PdfToImagesRequest(BaseModel):
    path: str
    pages: Optional[str] = None  # "1,3,5" or "1-3" or None for all
    dpi: int = 150
    output_folder: Optional[str] = None


@app.post("/pdf-to-images")
def pdf_to_images_endpoint(req: PdfToImagesRequest):
    """Render PDF pages as images."""
    import fitz

    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    doc = fitz.open(path)
    total = len(doc)

    # Parse pages
    if req.pages:
        page_nums = set()
        for part in req.pages.split(","):
            part = part.strip()
            if "-" in part:
                a, b = part.split("-", 1)
                page_nums.update(range(int(a), int(b) + 1))
            else:
                page_nums.add(int(part))
        zero_based = sorted(p - 1 for p in page_nums if 1 <= p <= total)
    else:
        _PAGE_CAP = 50  # prevent OOM on 500-page PDFs (specify pages param for more)
        zero_based = list(range(min(total, _PAGE_CAP)))

    if not zero_based:
        doc.close()
        raise HTTPException(status_code=400, detail=f"No valid pages. PDF has {total} pages.")

    out_folder = req.output_folder or os.path.join(os.path.dirname(path), os.path.splitext(os.path.basename(path))[0] + "_pages")
    os.makedirs(out_folder, exist_ok=True)

    zoom = req.dpi / 72
    mat = fitz.Matrix(zoom, zoom)
    saved = []
    for page_no in zero_based:
        page = doc[page_no]
        pix = page.get_pixmap(matrix=mat)
        img_path = os.path.join(out_folder, f"page_{page_no + 1}.png")
        pix.save(img_path)
        saved.append(img_path)

    doc.close()
    return {"success": True, "images": saved, "count": len(saved), "output_folder": out_folder}


class ImagesToPdfRequest(BaseModel):
    paths: list
    output_path: str


@app.post("/images-to-pdf")
def images_to_pdf_endpoint(req: ImagesToPdfRequest):
    """Combine images into a single PDF."""
    from PIL import Image

    if not req.paths:
        raise HTTPException(status_code=400, detail="No image paths provided")

    output = os.path.abspath(req.output_path)
    os.makedirs(os.path.dirname(output), exist_ok=True)

    images = []
    for p in req.paths:
        p = os.path.abspath(p)
        if not os.path.exists(p):
            raise HTTPException(status_code=404, detail=f"Image not found: {p}")
        img = Image.open(p)
        if img.mode == "RGBA":
            img = img.convert("RGB")
        images.append(img)

    if not images:
        raise HTTPException(status_code=400, detail="No valid images loaded")

    images[0].save(output, "PDF", save_all=True, append_images=images[1:])
    for img in images:
        img.close()

    return {"success": True, "path": output, "pages": len(images)}


# --- Image Tools (Segment 15B) ---

class ResizeImageRequest(BaseModel):
    path: str
    width: Optional[int] = None
    height: Optional[int] = None
    quality: int = 85
    output_path: Optional[str] = None


@app.post("/resize-image")
def resize_image_endpoint(req: ResizeImageRequest):
    """Resize or compress an image."""
    from PIL import Image
    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    img = Image.open(path)
    orig_size = img.size

    if req.width and req.height:
        img = img.resize((req.width, req.height), Image.LANCZOS)
    elif req.width:
        ratio = req.width / img.width
        img = img.resize((req.width, int(img.height * ratio)), Image.LANCZOS)
    elif req.height:
        ratio = req.height / img.height
        img = img.resize((int(img.width * ratio), req.height), Image.LANCZOS)

    output = os.path.abspath(req.output_path) if req.output_path else path
    os.makedirs(os.path.dirname(output), exist_ok=True)
    if img.mode == "RGBA" and output.lower().endswith((".jpg", ".jpeg")):
        img = img.convert("RGB")
    img.save(output, quality=req.quality)
    return {"success": True, "path": output, "original_size": list(orig_size), "new_size": list(img.size), "file_size": os.path.getsize(output)}


class ConvertImageRequest(BaseModel):
    path: str
    format: str  # jpg, png, webp, bmp
    output_path: Optional[str] = None
    quality: int = 90


@app.post("/convert-image")
def convert_image_endpoint(req: ConvertImageRequest):
    """Convert image to a different format."""
    from PIL import Image
    try:
        from pillow_heif import register_heif_opener
        register_heif_opener()
    except ImportError:
        pass

    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    fmt = req.format.lower().lstrip(".")
    ext_map = {"jpg": ".jpg", "jpeg": ".jpg", "png": ".png", "webp": ".webp", "bmp": ".bmp"}
    if fmt not in ext_map:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {fmt}. Use: jpg, png, webp, bmp")

    output = os.path.abspath(req.output_path) if req.output_path else os.path.splitext(path)[0] + ext_map[fmt]
    os.makedirs(os.path.dirname(output), exist_ok=True)

    img = Image.open(path)
    if img.mode == "RGBA" and fmt in ("jpg", "jpeg"):
        img = img.convert("RGB")
    img.save(output, quality=req.quality)
    return {"success": True, "path": output, "format": fmt, "size": os.path.getsize(output)}


class CropImageRequest(BaseModel):
    path: str
    x: int
    y: int
    width: int
    height: int
    output_path: Optional[str] = None


@app.post("/crop-image")
def crop_image_endpoint(req: CropImageRequest):
    """Crop an image to specified dimensions."""
    from PIL import Image
    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    img = Image.open(path)
    box = (req.x, req.y, req.x + req.width, req.y + req.height)
    cropped = img.crop(box)

    output = os.path.abspath(req.output_path) if req.output_path else path
    os.makedirs(os.path.dirname(output), exist_ok=True)
    cropped.save(output)
    return {"success": True, "path": output, "crop_box": list(box), "new_size": list(cropped.size)}


# --- Image Metadata (EXIF) ---

class ImageMetadataRequest(BaseModel):
    path: Optional[str] = None
    folder: Optional[str] = None


def _extract_exif(filepath: str) -> dict:
    """Extract EXIF metadata from a single image file."""
    from PIL import Image
    from PIL.ExifTags import IFD
    try:
        from pillow_heif import register_heif_opener
        register_heif_opener()
    except ImportError:
        pass

    img = Image.open(filepath)
    result = {
        "path": filepath,
        "dimensions": {"width": img.width, "height": img.height},
        "format": img.format or os.path.splitext(filepath)[1].lstrip(".").upper(),
    }

    exif_data = img.getexif()
    if not exif_data:
        img.close()
        result["exif"] = None
        return result

    def _rational(val):
        """Convert IFDRational or tuple to float."""
        if val is None:
            return None
        if hasattr(val, "numerator"):
            return float(val)
        if isinstance(val, tuple) and len(val) == 2:
            return val[0] / val[1] if val[1] else 0
        return float(val)

    exif = {}
    # DateTimeOriginal
    dt = exif_data.get(36867) or exif_data.get(306)  # DateTimeOriginal or DateTime
    if dt:
        exif["date_taken"] = str(dt).replace(":", "-", 2)  # 2025:02:20 → 2025-02-20

    # Camera
    make = (exif_data.get(271) or "").strip()
    model = (exif_data.get(272) or "").strip()
    if model:
        # Avoid "Canon Canon EOS 5D"
        exif["camera"] = model if make and model.startswith(make) else f"{make} {model}".strip()

    # Lens
    ifd_exif = exif_data.get_ifd(IFD.Exif)
    lens = ifd_exif.get(42036) if ifd_exif else None
    if lens:
        exif["lens"] = str(lens)

    # Focal length
    fl = ifd_exif.get(37386) if ifd_exif else None
    if fl is not None:
        v = _rational(fl)
        exif["focal_length"] = f"{v:.0f}mm" if v else None

    # Aperture
    fn = ifd_exif.get(33437) if ifd_exif else None
    if fn is not None:
        v = _rational(fn)
        exif["aperture"] = f"f/{v:.1f}" if v else None

    # Shutter speed
    et = ifd_exif.get(33434) if ifd_exif else None
    if et is not None:
        v = _rational(et)
        if v and v < 1:
            exif["shutter_speed"] = f"1/{int(round(1/v))}"
        elif v:
            exif["shutter_speed"] = f"{v:.1f}s"

    # ISO
    iso = ifd_exif.get(34855) if ifd_exif else None
    if iso is not None:
        exif["iso"] = int(iso) if not isinstance(iso, tuple) else int(iso[0])

    # GPS
    try:
        gps_ifd = exif_data.get_ifd(IFD.GPSInfo)
        if gps_ifd:
            def _dms_to_dd(dms, ref):
                d, m, s = [_rational(x) for x in dms]
                dd = d + m / 60 + s / 3600
                return -dd if ref in ("S", "W") else dd

            lat_dms = gps_ifd.get(2)
            lat_ref = gps_ifd.get(1)
            lon_dms = gps_ifd.get(4)
            lon_ref = gps_ifd.get(3)
            if lat_dms and lon_dms:
                exif["gps"] = {
                    "lat": round(_dms_to_dd(lat_dms, lat_ref), 6),
                    "lon": round(_dms_to_dd(lon_dms, lon_ref), 6),
                }
    except Exception:
        pass

    # Orientation
    orient = exif_data.get(274)
    if orient:
        exif["orientation"] = int(orient)

    img.close()
    result["exif"] = exif if exif else None
    return result


IMAGE_EXTS_EXIF = {".jpg", ".jpeg", ".png", ".tiff", ".tif", ".heic", ".heif", ".webp", ".bmp", ".dng", ".cr2", ".nef", ".arw"}


@app.post("/image-metadata")
def image_metadata_endpoint(req: ImageMetadataRequest):
    """Extract EXIF metadata from photos."""
    if req.path:
        path = os.path.abspath(req.path)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail=f"File not found: {path}")
        try:
            result = _extract_exif(path)
            result["_hint"] = "Metadata extracted. Answer the user's question about this image."
            return result
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read image: {e}")

    elif req.folder:
        folder = os.path.abspath(req.folder)
        if not os.path.isdir(folder):
            raise HTTPException(status_code=404, detail=f"Folder not found: {folder}")

        results = {}
        dates = []
        cameras = set()
        has_gps = 0
        total = 0

        for fname in sorted(os.listdir(folder)):
            if total >= 200:
                break
            ext = os.path.splitext(fname)[1].lower()
            if ext not in IMAGE_EXTS_EXIF:
                continue
            fpath = os.path.join(folder, fname)
            if not os.path.isfile(fpath):
                continue
            try:
                meta = _extract_exif(fpath)
                # Strip full path for batch — just filename
                meta["path"] = fname
                results[fname] = meta
                total += 1
                if meta.get("exif"):
                    if meta["exif"].get("date_taken"):
                        dates.append(meta["exif"]["date_taken"][:10])
                    if meta["exif"].get("camera"):
                        cameras.add(meta["exif"]["camera"])
                    if meta["exif"].get("gps"):
                        has_gps += 1
            except Exception:
                continue

        if not results:
            return {"folder": folder, "images_processed": 0, "results": {}, "_hint": "No images found in folder."}

        summary = {
            "date_range": f"{min(dates)} to {max(dates)}" if dates else None,
            "cameras": sorted(cameras) if cameras else [],
            "has_gps": has_gps,
            "total": total,
        }
        return {
            "folder": folder,
            "images_processed": total,
            "results": results,
            "summary": summary,
            "_hint": f"Metadata for {total} images. Summary shows date range and cameras used.",
        }

    else:
        raise HTTPException(status_code=400, detail="Provide path (single image) or folder (batch).")


# --- Archive Tools (Segment 15B) ---

class CompressFilesRequest(BaseModel):
    paths: list[str]
    output_path: str


@app.post("/compress-files")
def compress_files_endpoint(req: CompressFilesRequest):
    """Compress files into a zip archive."""
    import zipfile
    output = os.path.abspath(req.output_path)
    if not _is_safe_path(output):
        raise HTTPException(status_code=403, detail="Cannot write to system directories")
    os.makedirs(os.path.dirname(output), exist_ok=True)

    added = 0
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in req.paths:
            p = os.path.abspath(p)
            if not os.path.exists(p):
                continue
            if os.path.isdir(p):
                for root, dirs, files in os.walk(p):
                    for f in files:
                        fp = os.path.join(root, f)
                        arcname = os.path.relpath(fp, os.path.dirname(p))
                        zf.write(fp, arcname)
                        added += 1
            else:
                zf.write(p, os.path.basename(p))
                added += 1

    return {"success": True, "path": output, "files_added": added, "archive_size": os.path.getsize(output)}


class ExtractArchiveRequest(BaseModel):
    path: str
    output_path: Optional[str] = None


@app.post("/extract-archive")
def extract_archive_endpoint(req: ExtractArchiveRequest):
    """Extract a zip archive."""
    import zipfile
    path = os.path.abspath(req.path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    output = os.path.abspath(req.output_path) if req.output_path else os.path.splitext(path)[0]
    if not _is_safe_path(output):
        raise HTTPException(status_code=403, detail="Cannot extract to system directories")

    try:
        with zipfile.ZipFile(path, "r") as zf:
            # Security: check for path traversal (compute output abspath once)
            output_abs = os.path.abspath(output)
            for name in zf.namelist():
                target = os.path.join(output_abs, name)
                # normpath resolves .. without hitting disk (faster than abspath)
                if not os.path.normpath(target).startswith(output_abs):
                    raise HTTPException(status_code=400, detail=f"Unsafe path in archive: {name}")
            zf.extractall(output)
            return {"success": True, "path": output_abs, "files_extracted": len(zf.namelist())}
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Not a valid zip file")


# --- Download (Segment 15B) ---

class DownloadUrlRequest(BaseModel):
    url: str
    save_path: Optional[str] = None


@app.post("/download-url")
def download_url_endpoint(req: DownloadUrlRequest):
    """Download a file from a URL."""
    import urllib.request
    import urllib.error

    url = req.url.strip()
    if not url.startswith(("http://", "https://")):
        raise HTTPException(status_code=400, detail="URL must start with http:// or https://")

    # Determine save path
    if req.save_path:
        save_path = os.path.abspath(req.save_path)
    else:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        filename = os.path.basename(parsed.path) or "downloaded_file"
        save_dir = os.path.join(os.path.expanduser("~"), "Downloads", "Pinpoint")
        os.makedirs(save_dir, exist_ok=True)
        save_path = os.path.join(save_dir, filename)

    if not _is_safe_path(save_path):
        raise HTTPException(status_code=403, detail="Cannot save to system directories")
    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    try:
        req_obj = urllib.request.Request(url, headers={"User-Agent": "Pinpoint/1.0"})
        with urllib.request.urlopen(req_obj, timeout=60) as resp:
            with open(save_path, "wb") as f:
                shutil.copyfileobj(resp, f)  # stream in chunks, not buffer entire file in RAM
        return {"success": True, "path": save_path, "size": os.path.getsize(save_path), "url": url,
                "_hint": "Use index_file to make this file searchable, or send_file to share it."}
    except urllib.error.URLError as e:
        raise HTTPException(status_code=400, detail=f"Download failed: {e}")


# --- Smart Ops (Segment 15B) ---

class FindDuplicatesRequest(BaseModel):
    folder: str


@app.post("/find-duplicates")
def find_duplicates_endpoint(req: FindDuplicatesRequest):
    """Find duplicate files in a folder by content hash."""
    folder = os.path.abspath(req.folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    import hashlib
    hashes = {}  # hash → [paths]
    scanned = 0
    _MAX_FILES_SCAN = 20000  # cap to prevent multi-minute scans on huge folder trees
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for name in files:
            if scanned >= _MAX_FILES_SCAN:
                break
            fp = os.path.join(root, name)
            try:
                size = os.path.getsize(fp)
                if size == 0:
                    continue
                # Hash first 8KB + size for speed
                with open(fp, "rb") as f:
                    data = f.read(8192)
                key = f"{size}:{hashlib.md5(data).hexdigest()}"
                hashes.setdefault(key, []).append(fp)
                scanned += 1
            except OSError:
                continue
        if scanned >= _MAX_FILES_SCAN:
            break

    duplicates = {k: v for k, v in hashes.items() if len(v) > 1}
    groups = [{"files": paths, "count": len(paths)} for paths in duplicates.values()]
    groups.sort(key=lambda g: g["count"], reverse=True)

    resp = {
        "folder": folder,
        "scanned": scanned,
        "duplicate_groups": len(groups),
        "total_duplicates": sum(g["count"] - 1 for g in groups),
        "groups": groups[:50],  # cap at 50 groups
    }
    if groups:
        resp["_hint"] = "Use delete_file to remove duplicates (keep the original). Ask user which copies to remove."
    return resp


class BatchRenameRequest(BaseModel):
    folder: str
    pattern: str  # regex or glob pattern to match
    replace: str  # replacement string
    dry_run: bool = True  # preview only by default


@app.post("/batch-rename")
def batch_rename_endpoint(req: BatchRenameRequest):
    """Rename files in a folder matching a pattern. dry_run=true (default) shows preview only."""
    import re
    folder = os.path.abspath(req.folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")
    if not _is_safe_path(folder):
        raise HTTPException(status_code=403, detail="Cannot rename files in system directories")

    renamed = []
    errors = []
    try:
        regex = re.compile(req.pattern)
    except re.error as e:
        raise HTTPException(status_code=400, detail=f"Invalid regex pattern: {e}")

    for name in os.listdir(folder):
        new_name = regex.sub(req.replace, name)
        if new_name != name:
            old_path = os.path.join(folder, name)
            new_path = os.path.join(folder, new_name)
            if os.path.isfile(old_path) and not os.path.exists(new_path):
                if req.dry_run:
                    renamed.append({"old": name, "new": new_name})
                else:
                    try:
                        os.rename(old_path, new_path)
                        renamed.append({"old": name, "new": new_name})
                    except OSError as e:
                        errors.append({"file": name, "error": str(e)})

    resp = {"folder": folder, "renamed": len(renamed), "errors": len(errors), "details": renamed[:50], "dry_run": req.dry_run}
    if req.dry_run and renamed:
        resp["_hint"] = f"Preview: {len(renamed)} files would be renamed. Call again with dry_run=false to execute."
    return resp


# --- Run Python (Segment 15D) ---

PYTHON_WORK_DIR = "/tmp/pinpoint_python"
os.makedirs(PYTHON_WORK_DIR, exist_ok=True)


class RunPythonRequest(BaseModel):
    code: str
    timeout: int = 30


@app.post("/run-python")
def run_python_endpoint(req: RunPythonRequest):
    """Execute Python code and return stdout + created files."""
    import io
    import contextlib
    import signal

    timeout = min(req.timeout, 120)

    # Snapshot files before execution
    before = set()
    for root, dirs, files in os.walk(PYTHON_WORK_DIR):
        for f in files:
            before.add(os.path.join(root, f))

    # Pre-loaded namespace
    namespace = {
        "__builtins__": __builtins__,
        "WORK_DIR": PYTHON_WORK_DIR,
    }

    # Lazy imports inside namespace
    setup_code = """
import os, sys, json, math, re, pathlib, shutil, glob, hashlib, datetime, io, csv
WORK_DIR = "{work_dir}"
os.chdir(WORK_DIR)
try:
    import numpy as np
except ImportError:
    pass
try:
    import pandas as pd
except ImportError:
    pass
try:
    from PIL import Image, ImageDraw, ImageFont, ImageFilter
except ImportError:
    pass
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
except ImportError:
    pass
""".format(work_dir=PYTHON_WORK_DIR)

    full_code = setup_code + "\n" + req.code

    stdout_capture = io.StringIO()
    stderr_capture = io.StringIO()

    # Timeout handler
    def timeout_handler(signum, frame):
        raise TimeoutError(f"Code execution timed out after {timeout}s")

    try:
        old_handler = signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(timeout)

        with contextlib.redirect_stdout(stdout_capture), contextlib.redirect_stderr(stderr_capture):
            exec(full_code, namespace)

        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)
    except TimeoutError as e:
        signal.alarm(0)
        return {"success": False, "error": str(e), "stdout": stdout_capture.getvalue()[:5000]}
    except Exception as e:
        signal.alarm(0)
        return {"success": False, "error": f"{type(e).__name__}: {e}", "stdout": stdout_capture.getvalue()[:5000]}

    # Find new/modified files
    after = set()
    for root, dirs, files in os.walk(PYTHON_WORK_DIR):
        for f in files:
            after.add(os.path.join(root, f))
    new_files = sorted(after - before)

    stdout_text = stdout_capture.getvalue()[:10000]
    stderr_text = stderr_capture.getvalue()[:2000]

    result = {"success": True, "stdout": stdout_text}
    if stderr_text:
        result["stderr"] = stderr_text
    if new_files:
        result["files_created"] = new_files[:50]  # cap to prevent huge JSON responses
        if len(new_files) > 50:
            result["files_created_total"] = len(new_files)
    return result


# --- Persistent Memory (Segment 18F) ---

class MemorySaveRequest(BaseModel):
    fact: str
    category: str = "general"


def _memory_fts_search(conn, query: str, limit: int = 10, user_id: str = None) -> list:
    """Search memories using FTS5 BM25 (porter stemming). Falls back to LIKE if FTS fails."""
    try:
        # Build FTS query: extract words, quote each, join with OR for recall
        words = query.strip().split()
        stop = {"i", "my", "is", "am", "the", "a", "an", "to", "in", "on", "at", "for", "of", "and", "or", "that", "this", "it"}
        keywords = [w for w in words if w.lower() not in stop and len(w) > 1]
        if not keywords:
            keywords = [w for w in words if len(w) > 1][:3]
        if not keywords:
            return []
        fts_query = " OR ".join(f'"{k}"' for k in keywords)
        rows = conn.execute(
            """SELECT m.id, m.fact, m.category, bm25(memories_fts) AS rank
               FROM memories_fts f
               JOIN memories m ON m.id = f.rowid
               WHERE memories_fts MATCH ? AND m.superseded_by IS NULL
               ORDER BY rank
               LIMIT ?""",
            (fts_query, limit)
        ).fetchall()
        results = [{"id": r["id"], "fact": r["fact"], "category": r["category"]} for r in rows]
        if results:
            return results
    except Exception as e:
        print(f"[Memory] FTS search failed ({e}), falling back to LIKE")
    # Fallback: LIKE search
    candidates = {}
    for kw in keywords[:4]:
        rows = conn.execute(
            "SELECT id, fact, category FROM memories WHERE superseded_by IS NULL AND fact LIKE ? LIMIT 10",
            (f"%{kw}%",)
        ).fetchall()
        for row in rows:
            if row["id"] not in candidates:
                candidates[row["id"]] = {"id": row["id"], "fact": row["fact"], "category": row["category"]}
    return list(candidates.values())[:limit]


def _memory_find_similar(conn, fact: str, limit: int = 5) -> list:
    """Find existing memories similar to a new fact using FTS5 BM25."""
    return _memory_fts_search(conn, fact, limit)


def _memory_log_history(conn, memory_id: int, old_fact: str, new_fact: str, action: str):
    """Log a memory change to the audit trail."""
    now = datetime.utcnow().isoformat()
    conn.execute(
        "INSERT INTO memory_history(memory_id, old_fact, new_fact, action, created_at) VALUES (?, ?, ?, ?, ?)",
        (memory_id, old_fact, new_fact, action, now)
    )


def _memory_fts_sync(conn, memory_id: int, fact: str, delete_only: bool = False):
    """Keep memories_fts in sync with memories table."""
    try:
        conn.execute("DELETE FROM memories_fts WHERE rowid = ?", (memory_id,))
        if not delete_only:
            conn.execute("INSERT INTO memories_fts(rowid, fact) VALUES (?, ?)", (memory_id, fact))
    except Exception:
        pass  # FTS sync is best-effort


def _memory_decide_with_llm(new_fact: str, new_category: str, existing: list) -> dict:
    """Use Gemini flash-lite to decide how new fact relates to existing memories.
    Returns: {"action": "ADD|UPDATE|DELETE|NONE", "target_id": int|null, "merged_text": str|null}
    """
    GEMINI_KEY = os.getenv("GEMINI_API_KEY")
    if not GEMINI_KEY or not existing:
        return {"action": "ADD", "target_id": None, "merged_text": None}

    # Map real IDs to integers (anti-hallucination)
    id_map = {}
    display = []
    for i, m in enumerate(existing):
        id_map[str(i)] = m["id"]
        display.append(f'{{"id": {i}, "text": "{m["fact"]}", "category": "{m["category"]}"}}')

    prompt = f"""You manage a personal memory store. A new fact is being saved. Compare it to existing memories and decide what to do.

Existing memories:
[{", ".join(display)}]

New fact: "{new_fact}" (category: {new_category})

Decide ONE action:
- NONE: New fact is semantically the same as an existing memory. Don't save. Example: "Likes pizza" and "Loves pizza" = same meaning.
- ADD: New fact is genuinely new, not related to any existing memory.
- UPDATE: New fact is about the same subject but more specific/detailed. Replace the old memory. Example: "Likes hiking" → "Likes hiking in the Western Ghats on weekends".
- MERGE: New fact adds complementary info to an existing memory. Combine them. Example: "Likes cheese pizza" + "Likes chicken pizza" → "Likes cheese and chicken pizza".
- DELETE: New fact directly contradicts an existing memory. Remove old, save new. Example: "Loves pizza" → "Dislikes pizza".

Decide the best action."""

    _memory_decision_schema = {
        "type": "OBJECT",
        "properties": {
            "action": {"type": "STRING", "enum": ["ADD", "UPDATE", "MERGE", "DELETE", "NONE"]},
            "target_id": {"type": "INTEGER", "nullable": True},
            "merged_text": {"type": "STRING", "nullable": True},
        },
        "required": ["action"],
    }

    try:
        from google.genai import types as genai_types
        from extractors import _get_gemini, gemini_call_with_retry
        client = _get_gemini()
        if not client:
            return {"action": "ADD", "target_id": None, "merged_text": None}
        response = gemini_call_with_retry(
            client,
            model=os.environ.get("GEMINI_MODEL", "gemini-3.1-flash-lite-preview"),
            contents=prompt,
            config=genai_types.GenerateContentConfig(
                response_mime_type="application/json",
                response_json_schema=_memory_decision_schema,
            ),
        )

        import json as _json
        decision = _json.loads(response.text)

        # Map integer ID back to real DB ID
        if decision.get("target_id") is not None:
            real_id = id_map.get(str(decision["target_id"]))
            if real_id is None:
                return {"action": "ADD", "target_id": None, "merged_text": None}
            decision["target_id"] = real_id

        if decision.get("action") not in ("ADD", "UPDATE", "MERGE", "DELETE", "NONE"):
            return {"action": "ADD", "target_id": None, "merged_text": None}

        return decision
    except Exception as e:
        print(f"[Memory] LLM decision failed ({e}), falling back to ADD")
        return {"action": "ADD", "target_id": None, "merged_text": None}


@app.post("/memory")
def memory_save(req: MemorySaveRequest):
    """Save a personal fact to persistent memory. Uses LLM to detect duplicates, merge related facts, and handle contradictions."""
    if not req.fact.strip():
        raise HTTPException(status_code=400, detail="fact cannot be empty")
    conn = _get_conn()
    now = datetime.utcnow().isoformat()
    category = req.category.strip().lower()

    # Step 1: Find similar existing memories
    similar = _memory_find_similar(conn, req.fact)

    # Step 2: LLM decides what to do (or fallback to ADD if no similar)
    decision = _memory_decide_with_llm(req.fact.strip(), category, similar)
    action = decision.get("action", "ADD")
    target_id = decision.get("target_id")
    merged_text = decision.get("merged_text")

    result = {"success": True, "action": action}

    if action == "NONE":
        result["message"] = "Already remembered (semantically equivalent)"
        return result

    elif action == "UPDATE" and target_id:
        old_row = conn.execute("SELECT fact FROM memories WHERE id = ?", (target_id,)).fetchone()
        new_text = merged_text or req.fact.strip()
        conn.execute("UPDATE memories SET fact = ?, updated_at = ? WHERE id = ?", (new_text, now, target_id))
        _memory_log_history(conn, target_id, old_row["fact"] if old_row else None, new_text, "UPDATE")
        _memory_fts_sync(conn, target_id, new_text)
        conn.commit()
        result["id"] = target_id
        result["updated_text"] = new_text
        return result

    elif action == "MERGE" and target_id:
        old_row = conn.execute("SELECT fact FROM memories WHERE id = ?", (target_id,)).fetchone()
        new_text = merged_text or req.fact.strip()
        conn.execute("UPDATE memories SET fact = ?, updated_at = ? WHERE id = ?", (new_text, now, target_id))
        _memory_log_history(conn, target_id, old_row["fact"] if old_row else None, new_text, "MERGE")
        _memory_fts_sync(conn, target_id, new_text)
        conn.commit()
        result["id"] = target_id
        result["merged_text"] = new_text
        return result

    elif action == "DELETE" and target_id:
        old_row = conn.execute("SELECT fact FROM memories WHERE id = ?", (target_id,)).fetchone()
        conn.execute("UPDATE memories SET superseded_by = -1 WHERE id = ?", (target_id,))
        _memory_fts_sync(conn, target_id, "", delete_only=True)
        _memory_log_history(conn, target_id, old_row["fact"] if old_row else None, req.fact.strip(), "DELETE")
        cursor = conn.execute(
            "INSERT INTO memories(fact, category, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (req.fact.strip(), category, now, now)
        )
        new_id = cursor.lastrowid
        _memory_fts_sync(conn, new_id, req.fact.strip())
        _memory_log_history(conn, new_id, None, req.fact.strip(), "ADD")
        conn.commit()
        result["id"] = new_id
        result["superseded_id"] = target_id
        return result

    else:
        # ADD
        cursor = conn.execute(
            "INSERT INTO memories(fact, category, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (req.fact.strip(), category, now, now)
        )
        new_id = cursor.lastrowid
        _memory_fts_sync(conn, new_id, req.fact.strip())
        _memory_log_history(conn, new_id, None, req.fact.strip(), "ADD")
        conn.commit()
        result["id"] = new_id
        return result


@app.get("/memory/search")
def memory_search(
    q: str = Query(..., description="Search keywords"),
    limit: int = Query(10, ge=1, le=50),
):
    """Search persistent memories using FTS5 BM25."""
    conn = _get_conn()
    results = _memory_fts_search(conn, q, limit)
    return {"query": q, "count": len(results), "results": results}


@app.get("/memory/list")
def memory_list(
    category: Optional[str] = Query(None, description="Filter by category"),
    limit: int = Query(50, ge=1, le=200),
):
    """List all persistent memories (optionally filter by category)."""
    conn = _get_conn()
    if category:
        rows = conn.execute(
            "SELECT id, fact, category, created_at FROM memories WHERE category = ? ORDER BY updated_at DESC LIMIT ?",
            (category.lower(), limit)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, fact, category, created_at FROM memories ORDER BY updated_at DESC LIMIT ?",
            (limit,)
        ).fetchall()
    return {"count": len(rows), "memories": [dict(r) for r in rows]}


@app.delete("/memory/{memory_id}")
def memory_delete(memory_id: int):
    """Delete a memory by ID."""
    conn = _get_conn()
    cursor = conn.execute("DELETE FROM memories WHERE id = ?", (memory_id,))
    conn.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="Memory not found")
    return {"success": True, "deleted_id": memory_id}


class MemoryForgetRequest(BaseModel):
    description: str

@app.post("/memory/forget")
def memory_forget(req: MemoryForgetRequest):
    """Forget a memory by description — uses FTS5 to find best match and deletes it."""
    if not req.description.strip():
        raise HTTPException(status_code=400, detail="description cannot be empty")
    conn = _get_conn()
    # Use FTS5 to find matching memories
    candidates = _memory_fts_search(conn, req.description, limit=5)

    if not candidates:
        return {"success": False, "error": "No matching memory found", "searched_for": req.description}

    # Pick best match (FTS5 returns in BM25 rank order, first is best)
    best = candidates[0]
    _memory_log_history(conn, best["id"], best["fact"], None, "FORGET")
    _memory_fts_sync(conn, best["id"], "", delete_only=True)
    conn.execute("DELETE FROM memories WHERE id = ?", (best["id"],))
    conn.commit()
    return {"success": True, "deleted_id": best["id"], "deleted_fact": best["fact"]}


@app.get("/memory/context")
def memory_context(q: Optional[str] = Query(None, description="Current user message for query-relevant retrieval")):
    """Get active memories for system prompt. Static always included, dynamic filtered by relevance if query given."""
    conn = _get_conn()
    STATIC = ["preferences", "people", "places"]

    # Always include static memories (preferences/people/places)
    static_rows = conn.execute(
        "SELECT fact, category FROM memories WHERE superseded_by IS NULL AND category IN ('preferences', 'people', 'places') ORDER BY category, updated_at DESC"
    ).fetchall()

    lines = []
    groups = {}
    for r in static_rows:
        cat = r["category"] or "general"
        groups.setdefault(cat, []).append(r["fact"])
    for cat in STATIC:
        if cat in groups:
            lines.append(f"[{cat}]")
            for f in groups[cat]:
                lines.append(f"- {f}")
    static_count = len(static_rows)

    # Dynamic memories: query-relevant (FTS5) if query given, else all
    if q and q.strip():
        dynamic = _memory_fts_search(conn, q, limit=10)
        # Filter out static categories (already included above)
        dynamic = [d for d in dynamic if d.get("category", "general") not in STATIC]
        if dynamic:
            lines.append("[relevant]")
            for d in dynamic:
                lines.append(f"- {d['fact']}")
        total_count = static_count + len(dynamic)
    else:
        # No query — return all (backward compatible)
        dynamic_rows = conn.execute(
            "SELECT fact, category FROM memories WHERE superseded_by IS NULL AND category NOT IN ('preferences', 'people', 'places') ORDER BY category, updated_at DESC"
        ).fetchall()
        dyn_groups = {}
        for r in dynamic_rows:
            cat = r["category"] or "general"
            dyn_groups.setdefault(cat, []).append(r["fact"])
        for cat, facts in sorted(dyn_groups.items()):
            lines.append(f"[{cat}]")
            for f in facts:
                lines.append(f"- {f}")
        total_count = static_count + len(dynamic_rows)

    if not lines:
        return {"text": "", "count": 0}
    return {"text": "\n".join(lines), "count": total_count}


@app.get("/setting")
def setting_get(key: str = Query(...)):
    """Get a setting value."""
    conn = _get_conn()
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return {"key": key, "value": row["value"] if row else None}


@app.post("/setting")
def setting_set(key: str = Query(...), value: str = Query(...)):
    """Set a setting value."""
    conn = _get_conn()
    conn.execute(
        "INSERT INTO settings(key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value)
    )
    conn.commit()
    return {"success": True, "key": key, "value": value}


# --- Web Search (Segment 18G: Jina Reader) --- COMMENTED OUT (Segment 18U: replaced by web_read)
# @app.get("/web-search")
# def web_search(
#     q: str = Query(..., description="Search query"),
#     limit: int = Query(5, ge=1, le=10, description="Max results"),
# ):
#     """Search the web using Jina Reader API. Returns LLM-friendly text."""
#     import requests as req
#     try:
#         headers = {"Accept": "application/json"}
#         jina_key = os.environ.get("JINA_API_KEY")
#         if jina_key:
#             headers["Authorization"] = f"Bearer {jina_key}"
#         resp = req.get(
#             f"https://s.jina.ai/?q={q}",
#             headers=headers,
#             timeout=15,
#         )
#         resp.raise_for_status()
#         data = resp.json()
#         results = []
#         for item in (data.get("data", []) or [])[:limit]:
#             results.append({
#                 "title": item.get("title", ""),
#                 "url": item.get("url", ""),
#                 "content": (item.get("content") or "")[:1000],
#             })
#         return {"query": q, "count": len(results), "results": results}
#     except Exception as e:
#         return {"query": q, "count": 0, "results": [], "error": str(e)}


# --- Web Read (Segment 18U: browser tool, readability + html2text) ---

_WEB_READ_MAX = 8000  # max chars per chunk

# Lazy-loaded html2text converter
_h2t = None
def _get_html2text():
    global _h2t
    if _h2t is None:
        import html2text
        _h2t = html2text.HTML2Text()
        _h2t.ignore_images = True
        _h2t.body_width = 0       # no line wrapping
        _h2t.ignore_emphasis = False
        _h2t.protect_links = True
    return _h2t

def _paginate(text, start):
    """Paginate text: cut at nearest newline, return chunk + metadata."""
    import re
    text = re.sub(r"\n{3,}", "\n\n", text)
    total = len(text)
    end = start + _WEB_READ_MAX
    if end < total:
        nl = text.rfind("\n", start, end)
        if nl > start:
            end = nl + 1
    chunk = text[start:end]
    has_more = end < total
    if has_more:
        chunk += f"\n...(truncated — {total} total chars, call with start={end} for more)"
    return {"content": chunk, "total": total, "start": start, "end": end, "has_more": has_more}

@app.get("/web-read")
def web_read(
    url: str = Query(..., description="URL to fetch and read"),
    start: int = Query(0, ge=0, description="Character offset for pagination"),
):
    """Fetch any URL and return clean markdown text. Works like a browser."""
    import requests as req
    try:
        from readability import Document
    except ImportError:
        return {"url": url, "title": "", "content": "", "error": "Missing deps: pip install readability-lxml html2text"}
    try:
        resp = req.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate",
            },
            timeout=15,
            allow_redirects=True,
        )
        resp.raise_for_status()
        ct = resp.headers.get("content-type", "")
        # Non-HTML: return raw text (JSON, plain text, etc.)
        if "html" not in ct:
            result = _paginate(resp.text, start)
            result.update({"url": url, "title": ""})
            return result
        # Clean up search engine HTML before conversion
        from urllib.parse import unquote, urlparse, parse_qs
        import re
        html = resp.text
        # Decode DDG redirect URLs to actual URLs
        def _decode_ddg(m):
            try:
                params = parse_qs(urlparse(m.group(0)).query)
                return unquote(params.get("uddg", [m.group(0)])[0])
            except Exception:
                return m.group(0)
        html = re.sub(r'//duckduckgo\.com/l/\?[^"\'>\s]+', _decode_ddg, html)
        # Convert relative URLs to absolute (for Brave, Google, etc.)
        parsed = urlparse(url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        html = re.sub(r'href="/', f'href="{base}/', html)
        # Strip search engine nav/UI elements
        html = re.sub(r'<nav[^>]*>.*?</nav>', '', html, flags=re.DOTALL)
        html = re.sub(r'<header[^>]*>.*?</header>', '', html, flags=re.DOTALL)
        html = re.sub(r'<footer[^>]*>.*?</footer>', '', html, flags=re.DOTALL)
        # HTML → readability (extract main content) → html2text (convert to markdown)
        h = _get_html2text()
        doc = Document(html)
        title = doc.title() or ""
        md = h.handle(doc.summary())
        if len(md.strip()) < 200:
            # Readability failed (search results, etc.) — convert full page
            md = h.handle(html)
        # Remove lines that are just relative links or empty brackets
        md = re.sub(r'^\s*\[?\s*\]\(</[^)]*>\)\s*$', '', md, flags=re.MULTILINE)
        md = re.sub(r'\n{3,}', '\n\n', md)
        result = _paginate(md, start)
        result.update({"url": url, "title": title})
        return result
    except Exception as e:
        return {"url": url, "title": "", "content": "", "error": str(e)}


# --- Reminders (persistent) ---

class ReminderRequest(BaseModel):
    chat_jid: str
    message: str
    trigger_at: str  # ISO format
    repeat: Optional[str] = None  # daily, weekly, monthly, weekdays

@app.post("/reminders")
def save_reminder(req: ReminderRequest):
    """Save a reminder to the database."""
    conn = _get_conn()
    now = datetime.utcnow().isoformat()
    cur = conn.execute(
        "INSERT INTO reminders(chat_jid, message, trigger_at, repeat, created_at) VALUES (?, ?, ?, ?, ?)",
        (req.chat_jid, req.message, req.trigger_at, req.repeat, now)
    )
    conn.commit()
    return {"id": cur.lastrowid, "message": req.message, "trigger_at": req.trigger_at, "repeat": req.repeat}

@app.get("/reminders")
def list_reminders_endpoint(chat_jid: Optional[str] = Query(None)):
    """Load all reminders (optionally filtered by chat_jid)."""
    conn = _get_conn()
    if chat_jid:
        rows = conn.execute("SELECT * FROM reminders WHERE chat_jid = ? ORDER BY trigger_at", (chat_jid,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM reminders ORDER BY trigger_at").fetchall()
    return {"reminders": [dict(r) for r in rows]}

@app.delete("/reminders/{reminder_id}")
def delete_reminder(reminder_id: int):
    """Delete a reminder."""
    conn = _get_conn()
    conn.execute("DELETE FROM reminders WHERE id = ?", (reminder_id,))
    conn.commit()
    return {"success": True, "id": reminder_id}

@app.put("/reminders/{reminder_id}")
def update_reminder(reminder_id: int, trigger_at: str = Query(...)):
    """Update a reminder's trigger time (used for rescheduling recurring reminders)."""
    conn = _get_conn()
    conn.execute("UPDATE reminders SET trigger_at = ? WHERE id = ?", (trigger_at, reminder_id))
    conn.commit()
    return {"success": True, "id": reminder_id, "trigger_at": trigger_at}


# --- Table Extraction from PDFs (pdfplumber) ---

@app.post("/extract-tables")
def extract_tables_endpoint(
    path: str = Query(..., description="Path to PDF file"),
    pages: Optional[str] = Query(None, description="Page range: '1-5', '3', 'all'. Default: all"),
):
    """Extract structured tables from a PDF using pdfplumber."""
    path = os.path.abspath(path)
    if not os.path.exists(path):
        raise HTTPException(404, f"File not found: {path}")
    if not path.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files supported")

    import pdfplumber

    try:
        pdf = pdfplumber.open(path)
    except Exception as e:
        return {"error": f"Cannot open PDF: {e}"}

    total_pages = len(pdf.pages)

    # Parse page range
    page_indices = []
    if not pages or pages == "all":
        page_indices = list(range(total_pages))
    elif "-" in pages:
        parts = pages.split("-")
        start = max(int(parts[0]) - 1, 0)
        end = min(int(parts[1]), total_pages)
        page_indices = list(range(start, end))
    else:
        p = int(pages) - 1
        if 0 <= p < total_pages:
            page_indices = [p]

    tables = []
    for pi in page_indices:
        page = pdf.pages[pi]
        page_tables = page.extract_tables()
        for ti, table in enumerate(page_tables):
            if not table or len(table) < 2:
                continue
            # First row as headers, rest as data
            headers = [str(c).strip() if c else "" for c in table[0]]
            rows = []
            for row in table[1:]:
                rows.append([str(c).strip() if c else "" for c in row])
            tables.append({
                "page": pi + 1,
                "table_index": ti + 1,
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
            })

    pdf.close()

    if not tables:
        return {"path": path, "total_pages": total_pages, "tables": [],
                "error": "No tables found in this PDF. Tables must be structured (not scanned images).",
                "_hint": "No tables detected. If the PDF is scanned, try OCR first with ocr(path)."}

    return {
        "path": path,
        "total_pages": total_pages,
        "tables_found": len(tables),
        "tables": tables,
        "_hint": "Tables extracted. Present them clearly or use analyze_data for further analysis.",
    }


# --- Folder Watcher (auto-index new/modified files) ---

_watchers = {}  # folder → Observer
_watch_debounce = {}  # filepath → Timer

# Supported file extensions for auto-indexing (same as extractors.py)
_WATCH_EXTS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
    ".txt", ".csv", ".epub", ".md", ".json", ".log",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp", ".heic", ".heif",
}


def _auto_index_file(path: str):
    """Auto-index a single file (called by watcher after debounce)."""
    _watch_debounce.pop(path, None)
    if not os.path.exists(path):
        return
    ext = os.path.splitext(path)[1].lower()
    if ext not in _WATCH_EXTS:
        return
    try:
        from extractors import extract
        from database import upsert_document, chunk_document
        result = extract(path)
        if result is None:
            return
        conn = _get_conn()
        upsert_document(conn, path, result["text"], result["file_type"], result.get("page_count", 0))
        doc_row = conn.execute("SELECT id FROM documents WHERE path = ?", (path,)).fetchone()
        if doc_row:
            try:
                chunk_document(conn, doc_row["id"], result["text"])
            except Exception:
                pass
        print(f"[Watcher] Auto-indexed: {os.path.basename(path)}")
    except Exception as e:
        print(f"[Watcher] Failed to index {path}: {e}")


def _start_watcher(folder: str):
    """Start watching a folder for file changes."""
    import threading
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler

    class IndexHandler(FileSystemEventHandler):
        def _handle(self, event):
            if event.is_directory:
                return
            path = event.src_path
            ext = os.path.splitext(path)[1].lower()
            if ext not in _WATCH_EXTS:
                return
            # Debounce: wait 5s after last change before indexing
            old = _watch_debounce.pop(path, None)
            if old:
                old.cancel()
            t = threading.Timer(5.0, _auto_index_file, args=[path])
            _watch_debounce[path] = t
            t.start()

        def on_created(self, event):
            self._handle(event)

        def on_modified(self, event):
            self._handle(event)

    observer = Observer()
    observer.schedule(IndexHandler(), folder, recursive=True)
    observer.daemon = True
    observer.start()
    _watchers[folder] = observer
    print(f"[Watcher] Watching: {folder}")


def _stop_watcher(folder: str):
    """Stop watching a folder."""
    obs = _watchers.pop(folder, None)
    if obs:
        obs.stop()
        print(f"[Watcher] Stopped: {folder}")


@app.post("/watch")
def watch_folder_endpoint(folder: str = Query(..., description="Folder to watch")):
    """Start auto-indexing new/modified files in a folder."""
    folder = os.path.abspath(folder)
    if not os.path.isdir(folder):
        raise HTTPException(400, f"Not a directory: {folder}")
    if folder in _watchers:
        return {"success": True, "folder": folder, "note": "Already watching"}
    _start_watcher(folder)
    # Persist to settings
    conn = _get_conn()
    row = conn.execute("SELECT value FROM settings WHERE key = 'watched_folders'").fetchone()
    watched = set(row["value"].split("|")) if row and row["value"] else set()
    watched.add(folder)
    conn.execute(
        "INSERT INTO settings(key, value) VALUES ('watched_folders', ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        ("|".join(watched),)
    )
    conn.commit()
    return {"success": True, "folder": folder, "_hint": f"Now watching {folder}. New/modified files will be auto-indexed."}


@app.post("/unwatch")
def unwatch_folder_endpoint(folder: str = Query(..., description="Folder to stop watching")):
    """Stop auto-indexing a folder."""
    folder = os.path.abspath(folder)
    if folder not in _watchers:
        return {"success": False, "error": f"Not watching: {folder}"}
    _stop_watcher(folder)
    # Remove from settings
    conn = _get_conn()
    row = conn.execute("SELECT value FROM settings WHERE key = 'watched_folders'").fetchone()
    if row and row["value"]:
        watched = set(row["value"].split("|"))
        watched.discard(folder)
        conn.execute(
            "INSERT INTO settings(key, value) VALUES ('watched_folders', ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            ("|".join(watched) if watched else "",)
        )
        conn.commit()
    return {"success": True, "folder": folder, "_hint": f"Stopped watching {folder}."}


@app.get("/watched")
def list_watched():
    """List all currently watched folders."""
    return {"folders": list(_watchers.keys()), "count": len(_watchers)}


# Restore watchers on startup
@app.on_event("startup")
def _restore_watchers():
    try:
        conn = _get_conn()
        row = conn.execute("SELECT value FROM settings WHERE key = 'watched_folders'").fetchone()
        if row and row["value"]:
            for folder in row["value"].split("|"):
                if folder and os.path.isdir(folder):
                    _start_watcher(folder)
    except Exception as e:
        print(f"[Watcher] Failed to restore watchers: {e}")


# --- Photo Cull (Segment 21) ---

class ScorePhotoRequest(BaseModel):
    path: str

class CullPhotosRequest(BaseModel):
    folder: str
    keep_pct: int = 80
    rejects_folder: Optional[str] = None

@app.post("/score-photo")
def api_score_photo(req: ScorePhotoRequest):
    """Score a photo's technical + aesthetic quality (Gemini vision, /100)."""
    from photo_cull import score_photo
    return score_photo(req.path)

@app.post("/cull-photos")
def api_cull_photos(req: CullPhotosRequest):
    """Auto-cull photos: score all, move bottom rejects to _rejects folder. Background job."""
    from photo_cull import cull_photos
    return cull_photos(req.folder, req.keep_pct, req.rejects_folder)

@app.get("/cull-photos/status")
def api_cull_status(folder: str = Query(..., description="Folder being culled"), cancel: bool = Query(False, description="Set true to cancel the job")):
    """Poll cull job progress. Set cancel=true to stop."""
    from photo_cull import get_cull_status
    return get_cull_status(folder, cancel=cancel)


# --- Photo Group (Segment 21B) ---

class SuggestCategoriesRequest(BaseModel):
    folder: str

@app.post("/suggest-categories")
def api_suggest_categories(req: SuggestCategoriesRequest):
    """Sample photos and suggest grouping categories via Gemini vision."""
    from photo_cull import suggest_categories
    return suggest_categories(req.folder)

class GroupPhotosRequest(BaseModel):
    folder: str
    categories: list
    uncategorized_folder: Optional[str] = None

@app.post("/group-photos")
def api_group_photos(req: GroupPhotosRequest):
    """Auto-group photos: classify ALL images via Gemini vision, move to category subfolders. Background job."""
    from photo_cull import group_photos
    return group_photos(req.folder, req.categories, req.uncategorized_folder)

@app.get("/group-photos/status")
def api_group_status(folder: str = Query(..., description="Folder being grouped"), cancel: bool = Query(False, description="Set true to cancel the job")):
    """Poll group job progress. Set cancel=true to stop."""
    from photo_cull import get_group_status
    return get_group_status(folder, cancel=cancel)


# --- Run ---

if __name__ == "__main__":
    import uvicorn
    print("[Pinpoint API] Starting on http://localhost:5123")
    print("[Pinpoint API] Docs at http://localhost:5123/docs")
    uvicorn.run(app, host="0.0.0.0", port=5123, log_level="info")
