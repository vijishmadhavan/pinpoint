"""
Pinpoint — Document extractors (tiered, speed-optimized)

Digital PDF:         PyMuPDF4LLM  (instant, CPU)
Scanned PDF:         Tesseract (0.5s/page CPU, 125 languages, auto script detection)
DOCX/XLSX/PPTX/EPUB: MarkItDown  (instant, CPU)
TXT/CSV/LOG/MD:      direct read  (instant)
Images:              Gemini 3.1 Flash-Lite captioning (multimodal, cheap)
"""

from __future__ import annotations

import os
import time
from typing import Any

# Max dimension for image preprocessing (originals untouched, resize in memory)
MAX_IMAGE_DIM = 1024

# OCR DPI: 200 for batch/indexing (fast), 300 for single-file (quality)
OCR_DPI = int(os.environ.get("OCR_DPI", "200"))

# Tesseract availability detection
_HAS_TESSERACT = False
try:
    import pytesseract

    pytesseract.get_tesseract_version()
    _HAS_TESSERACT = True
except Exception:
    pass
print(f"[Pinpoint] Tesseract: {'available' if _HAS_TESSERACT else 'not installed — Gemini OCR fallback'}")

# Minimum text per page to consider it "digital" (not scanned)
_MIN_TEXT_PER_PAGE = 50


# Script → Tesseract lang mapping for common scripts
_SCRIPT_TO_LANG = {
    "Latin": "eng",
    "Devanagari": "hin",
    "Malayalam": "mal",
    "Tamil": "tam",
    "Telugu": "tel",
    "Kannada": "kan",
    "Bengali": "ben",
    "Gujarati": "guj",
    "Gurmukhi": "pan",
    "Arabic": "ara",
    "Cyrillic": "rus",
    "Han": "chi_sim",
    "Hangul": "kor",
    "Japanese": "jpn",
    "Thai": "tha",
}


def _detect_tesseract_lang(img: Any) -> str:
    """Detect script from image and return best Tesseract lang code."""
    import pytesseract

    try:
        osd = pytesseract.image_to_osd(img)
        for line in osd.split("\n"):
            if line.startswith("Script:"):
                script = line.split(":")[1].strip()
                return _SCRIPT_TO_LANG.get(script, "eng")
    except Exception:
        pass
    return "eng"


def _ocr_tesseract(images: list[Any]) -> str:
    """OCR a list of PIL images using Tesseract with auto language detection."""
    import pytesseract

    texts = []
    for img in images:
        lang = _detect_tesseract_lang(img)
        text = pytesseract.image_to_string(img, lang=lang)
        if text.strip():
            texts.append(text.strip())

    return "\n\n".join(texts)


def _ocr_gemini(images: list[Any]) -> str:
    """OCR via Gemini vision when Tesseract unavailable."""
    client = _get_gemini()
    if not client:
        return ""
    import io

    from google.genai import types

    texts = []
    for img in images:
        try:
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            response = gemini_call_with_retry(
                client,
                model=os.environ.get("GEMINI_MODEL", "gemini-3.1-flash-lite-preview"),
                contents=[
                    types.Content(
                        parts=[
                            types.Part.from_bytes(data=buf.getvalue(), mime_type="image/jpeg"),
                            types.Part.from_text(
                                text="Extract ALL text from this image exactly as written. Preserve formatting and line breaks. Return ONLY the extracted text."
                            ),
                        ]
                    )
                ],
                config=types.GenerateContentConfig(
                    media_resolution=types.MediaResolution.MEDIA_RESOLUTION_LOW,
                ),
            )
            text = (response.text or "").strip()
            if text:
                texts.append(text)
        except Exception as e:
            print(f"[Gemini OCR] Failed on image: {e}")
    return "\n\n".join(texts)


def _preprocess_image(img: Any, max_dim: int = MAX_IMAGE_DIM) -> Any:
    """Resize image if larger than max_dim. Returns new image (original untouched)."""
    from PIL import Image

    w, h = img.size
    if max(w, h) <= max_dim:
        return img
    scale = max_dim / max(w, h)
    new_w, new_h = int(w * scale), int(h * scale)
    return img.resize((new_w, new_h), Image.LANCZOS)


def _is_scanned_pdf(doc: Any) -> bool:
    """Check if a PDF is scanned (no extractable text) vs digital.
    Samples up to 3 pages. If most have little text → scanned."""
    pages_to_check = min(len(doc), 3)
    empty_pages = 0
    for i in range(pages_to_check):
        text = doc[i].get_text().strip()
        if len(text) < _MIN_TEXT_PER_PAGE:
            empty_pages += 1
    return empty_pages > pages_to_check / 2


def extract_pdf(path: str) -> dict | None:
    """
    Extract text from a PDF using tiered approach:
    - Digital PDF → PyMuPDF4LLM (instant, CPU, markdown with tables)
    - Scanned PDF → Tesseract OCR (0.5s/page, 125 languages)
    """
    if not os.path.exists(path):
        print(f"[SKIP] File not found: {path}")
        return None

    import fitz

    try:
        doc = fitz.open(path)
        try:
            if doc.is_encrypted:
                print(f"[SKIP] Password-protected PDF: {path}")
                return None
            page_count = len(doc)
            scanned = _is_scanned_pdf(doc)
        finally:
            doc.close()
    except Exception as e:
        print(f"[SKIP] Cannot open PDF: {path} — {e}")
        return None

    if scanned:
        print(f"[OCR] Scanned PDF detected, using Tesseract: {path}")
        try:
            import fitz as fitz_render
            from PIL import Image

            doc_render = fitz_render.open(path)
            try:
                total_pages = len(doc_render)
                _PAGE_BATCH = 20  # process 20 pages at a time (~200MB RAM max vs 2GB+ for 200 pages)
                all_texts = []
                for batch_start in range(0, total_pages, _PAGE_BATCH):
                    batch_end = min(batch_start + _PAGE_BATCH, total_pages)
                    images = []
                    for page_num in range(batch_start, batch_end):
                        pix = doc_render[page_num].get_pixmap(dpi=OCR_DPI)
                        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                        images.append(img)
                    batch_text = (
                        _ocr_gemini(images) if _get_gemini() else (_ocr_tesseract(images) if _HAS_TESSERACT else "")
                    )
                    if batch_text.strip():
                        all_texts.append(batch_text.strip())
                    del images  # free batch memory before next batch
            finally:
                doc_render.close()
            text = "\n\n".join(all_texts)
        except Exception as e:
            print(f"[SKIP] OCR failed: {path} — {e}")
            return None
    else:
        # Digital PDF → PyMuPDF4LLM (instant, CPU)
        try:
            import pymupdf4llm

            text = pymupdf4llm.to_markdown(path)
        except Exception as e:
            print(f"[SKIP] PyMuPDF4LLM failed: {path} — {e}")
            return None

    return {
        "path": os.path.abspath(path),
        "text": text.strip(),
        "page_count": page_count,
        "file_type": "pdf",
    }


_MAX_EXCEL_ROWS_INDEX = 5000  # cap rows per sheet for indexing (prevents OOM on 100MB files)


def _extract_excel_all_sheets(path: str) -> str | None:
    """Extract ALL sheets from an Excel file as markdown tables (capped per sheet)."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= _MAX_EXCEL_ROWS_INDEX:
                    break
                rows.append([str(v) if v is not None else "" for v in row])
            if not rows:
                continue
            total_rows = ws.max_row or len(rows)
            # Build markdown table
            header = f"## Sheet: {sheet_name} ({total_rows} rows)"
            if total_rows > _MAX_EXCEL_ROWS_INDEX:
                header += f" [showing first {_MAX_EXCEL_ROWS_INDEX}]"
            lines = [header + "\n"]
            lines.append("| " + " | ".join(rows[0]) + " |")
            lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
            for row in rows[1:]:
                padded = row + [""] * (len(rows[0]) - len(row))
                lines.append("| " + " | ".join(padded[: len(rows[0])]) + " |")
            parts.append("\n".join(lines))
        wb.close()
        return "\n\n".join(parts) if parts else None
    except Exception as e:
        print(f"[SKIP] Excel multi-sheet extraction failed: {path} — {e}")
        return None


def extract_office(path: str) -> dict | None:
    """Extract text from DOCX/XLSX/PPTX/EPUB using Microsoft MarkItDown.
    For Excel (.xlsx/.xlsm): reads ALL sheets as markdown tables."""
    if not os.path.exists(path):
        print(f"[SKIP] File not found: {path}")
        return None

    ext = os.path.splitext(path)[1].lower()

    # Excel: use openpyxl to read ALL sheets (MarkItDown only reads active sheet)
    if ext in (".xlsx", ".xlsm"):
        text = _extract_excel_all_sheets(path)
        if text is None:
            # Fallback to MarkItDown
            try:
                from markitdown import MarkItDown

                md = MarkItDown(enable_plugins=False)
                result = md.convert(path)
                text = result.text_content
            except Exception as e:
                print(f"[SKIP] MarkItDown failed for {ext.upper()}: {path} — {e}")
                return None
        return {
            "path": os.path.abspath(path),
            "text": text.strip(),
            "page_count": 0,
            "file_type": ext.lstrip("."),
        }

    # Non-Excel office docs: use MarkItDown
    try:
        from markitdown import MarkItDown

        md = MarkItDown(enable_plugins=False)
        result = md.convert(path)
        text = result.text_content
    except Exception as e:
        print(f"[SKIP] MarkItDown failed for {ext.upper()}: {path} — {e}")
        return None

    file_type = ext.lstrip(".")
    return {
        "path": os.path.abspath(path),
        "text": text.strip(),
        "page_count": 0,
        "file_type": file_type,
    }


# --- Gemini captioning ---

_gemini_client = None


def _get_gemini() -> Any | None:
    """Lazy-load Gemini client for image captioning."""
    global _gemini_client
    if _gemini_client is None:
        from google import genai

        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            return None
        _gemini_client = genai.Client(api_key=api_key)
    return _gemini_client


def gemini_call_with_retry(client: Any, model: str, contents: Any, config: Any | None = None, retries: int = 2) -> Any:
    """Call Gemini with exponential backoff on 429/503 transient errors."""
    for attempt in range(retries + 1):
        try:
            kwargs = {"model": model, "contents": contents}
            if config:
                kwargs["config"] = config
            return client.models.generate_content(**kwargs)
        except Exception as e:
            err_str = str(e)
            is_transient = "429" in err_str or "503" in err_str or "RESOURCE_EXHAUSTED" in err_str
            if is_transient and attempt < retries:
                wait = 2 ** (attempt + 1)  # 2s, 4s
                print(f"[Gemini] Transient error ({err_str[:60]}), retry in {wait}s...")
                time.sleep(wait)
                continue
            raise


def extract_image(path: str) -> dict | None:
    """Caption an image using Gemini Flash-Lite. Requires GEMINI_API_KEY in .env."""
    if not os.path.exists(path):
        print(f"[SKIP] File not found: {path}")
        return None

    client = _get_gemini()
    if not client:
        print(f"[SKIP] GEMINI_API_KEY not set: {path}")
        return None

    try:
        import io

        from PIL import Image

        img = Image.open(path).convert("RGB")
        img = _preprocess_image(img, 384)  # Gemini LOW res = 384px
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=80)

        from google.genai import types

        response = gemini_call_with_retry(
            client,
            model=os.environ.get("GEMINI_MODEL", "gemini-3.1-flash-lite-preview"),
            contents=[
                types.Content(
                    parts=[
                        types.Part.from_bytes(data=buf.getvalue(), mime_type="image/jpeg"),
                        types.Part.from_text(
                            text="Describe this image in 1-2 sentences. Be specific about objects, people, text, and scene."
                        ),
                    ]
                ),
            ],
            config=types.GenerateContentConfig(
                media_resolution=types.MediaResolution.MEDIA_RESOLUTION_LOW,
            ),
        )
        caption = response.text.strip()
        if not caption:
            print(f"[SKIP] Gemini returned empty caption: {path}")
            return None

        return {
            "path": os.path.abspath(path),
            "text": caption,
            "page_count": 0,
            "file_type": "image",
        }
    except Exception as e:
        print(f"[SKIP] Gemini captioning failed: {path} — {e}")
        return None


# Extension routing
OFFICE_EXTENSIONS = {".docx", ".xlsx", ".pptx", ".epub"}
# NOT supported: .doc, .xls, .ppt (old binary formats — need LibreOffice)

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif", ".heic"}

TEXT_EXTENSIONS = {".txt", ".csv", ".log", ".md"}


_MAX_TEXT_SIZE = 50 * 1024 * 1024  # 50MB limit for text files


def extract_text(path: str) -> dict | None:
    """Read a plain text or CSV file. Handles encoding fallback."""
    if not os.path.exists(path):
        print(f"[SKIP] File not found: {path}")
        return None

    if os.path.getsize(path) > _MAX_TEXT_SIZE:
        print(f"[SKIP] File too large ({os.path.getsize(path)} bytes): {path}")
        return None

    ext = os.path.splitext(path)[1].lower()

    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            with open(path, encoding=encoding) as f:
                text = f.read()
            break
        except UnicodeDecodeError:
            continue
    else:
        print(f"[SKIP] Cannot decode file: {path}")
        return None

    file_type = "csv" if ext == ".csv" else "txt"

    return {
        "path": os.path.abspath(path),
        "text": text.strip(),
        "page_count": 0,
        "file_type": file_type,
    }


def extract(path: str) -> dict | None:
    """Universal extractor — picks the right method based on file extension."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".pdf":
        return extract_pdf(path)

    if ext in OFFICE_EXTENSIONS:
        return extract_office(path)

    if ext in IMAGE_EXTENSIONS:
        return extract_image(path)

    if ext in TEXT_EXTENSIONS:
        return extract_text(path)

    print(f"[SKIP] Unsupported file type: {ext} — {path}")
    return None


# --- Quick test ---
if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        result = extract(sys.argv[1])
        if result:
            print(f"Path:        {result['path']}")
            print(f"Pages:       {result['page_count']}")
            print(f"File type:   {result['file_type']}")
            print(f"Text length: {len(result['text'])} chars")
            print("--- First 1000 chars ---")
            print(result["text"][:1000])
        else:
            print("Failed to extract.")
    else:
        print("Usage: python extractors.py <path_to_file>")
