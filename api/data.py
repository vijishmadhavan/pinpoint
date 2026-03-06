"""Data analysis & calculation endpoints — /calculate, /read_excel, /analyze-data, /extract-tables, /pdf-to-excel."""

from __future__ import annotations

import ast
import math
import os
import time
from typing import Any

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel

from api.helpers import _check_safe, record_generated_file

router = APIRouter()


# --- Calculate (safe math) ---


class CalculateRequest(BaseModel):
    expression: str


# Safe functions/names for math evaluation
_SAFE_MATH = {
    "abs": abs,
    "round": round,
    "min": min,
    "max": max,
    "sum": sum,
    "len": len,
    "int": int,
    "float": float,
    "pi": math.pi,
    "e": math.e,
    "sqrt": math.sqrt,
    "ceil": math.ceil,
    "floor": math.floor,
    "log": math.log,
    "log10": math.log10,
    "pow": math.pow,
}


def _safe_eval(expression: str) -> float | int:
    """Safely evaluate a math expression using AST parsing."""
    # Parse the expression into an AST
    tree = ast.parse(expression, mode="eval")

    # Walk the AST and check all nodes are safe
    for node in ast.walk(tree):
        if isinstance(
            node,
            (
                ast.Expression,
                ast.BinOp,
                ast.UnaryOp,
                ast.Constant,
                ast.Add,
                ast.Sub,
                ast.Mult,
                ast.Div,
                ast.FloorDiv,
                ast.Mod,
                ast.Pow,
                ast.USub,
                ast.UAdd,
                ast.Call,
                ast.Name,
                ast.Load,
                ast.List,
                ast.Tuple,
            ),
        ):
            continue
        raise ValueError(f"Unsupported operation: {type(node).__name__}")

    # Compile and evaluate with only safe builtins
    code = compile(tree, "<calc>", "eval")
    return eval(code, {"__builtins__": {}}, _SAFE_MATH)


@router.post("/calculate")
def calculate_endpoint(req: CalculateRequest) -> dict:
    """Safely evaluate a mathematical expression."""
    expr = req.expression.strip()
    if not expr:
        raise HTTPException(status_code=400, detail="Empty expression")

    try:
        result = _safe_eval(expr)
        # Format with commas for large numbers
        if isinstance(result, float) and result == int(result) and abs(result) < 1e15:
            formatted = f"{int(result):,}"
        elif isinstance(result, (int,)):
            formatted = f"{result:,}"
        else:
            formatted = f"{result:,.4f}".rstrip("0").rstrip(".")
        return {"expression": expr, "result": result, "formatted": formatted}
    except (ValueError, SyntaxError, TypeError, ZeroDivisionError) as e:
        raise HTTPException(status_code=400, detail=f"Invalid expression: {e}")


# --- Read Excel ---


class ReadExcelRequest(BaseModel):
    path: str
    sheet_name: str | None = None
    cell_range: str | None = None


@router.post("/read_excel")
def read_excel_endpoint(req: ReadExcelRequest) -> dict:
    """Read specific cells/ranges from an Excel file via openpyxl."""
    path = os.path.abspath(req.path)
    _check_safe(path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    if not path.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm files supported")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot open Excel file: {e}")

    try:
        sheet_names = wb.sheetnames
        result = {"path": path, "sheet_names": sheet_names}

        # Select sheet
        if req.sheet_name:
            if req.sheet_name not in sheet_names:
                raise HTTPException(status_code=400, detail=f"Sheet '{req.sheet_name}' not found. Available: {sheet_names}")
            ws = wb[req.sheet_name]
        else:
            ws = wb.active

        result["active_sheet"] = ws.title

        if req.cell_range:
            # Read specific range
            try:
                cells = ws[req.cell_range]
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid range '{req.cell_range}': {e}")

            # Handle single cell
            if not isinstance(cells, tuple):
                result["data"] = [[str(cells.value) if cells.value is not None else ""]]
                result["rows"] = 1
                result["cols"] = 1
            else:
                # Range of cells — could be rows of tuples or a single tuple
                rows = []
                if isinstance(cells[0], tuple):
                    for row in cells:
                        rows.append([str(c.value) if c.value is not None else "" for c in row])
                else:
                    # Single row/column
                    rows.append([str(c.value) if c.value is not None else "" for c in cells])
                result["data"] = rows
                result["rows"] = len(rows)
                result["cols"] = len(rows[0]) if rows else 0
        else:
            # No range — return first 20 rows
            rows = []
            for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
                rows.append([str(v) if v is not None else "" for v in row])
            result["data"] = rows
            result["rows"] = len(rows)
            result["cols"] = len(rows[0]) if rows else 0

        # Format as markdown table for readability
        if result.get("data") and len(result["data"]) > 0:
            lines = []
            for i, row in enumerate(result["data"]):
                lines.append("| " + " | ".join(row) + " |")
                if i == 0:
                    lines.append("| " + " | ".join(["---"] * len(row)) + " |")
            result["table"] = "\n".join(lines)

        return result
    finally:
        wb.close()


# --- Pandas data analysis (Segment 15 + 18Q: smart cache, multi-sheet, search) ---

# DataFrame LRU cache: (path, mtime, sheet) → DataFrame
import threading

_df_cache = {}  # key → {"df": DataFrame, "atime": float}
_df_cache_lock = threading.Lock()
_DF_CACHE_MAX = 5


def _normalize_for_search(val: str) -> str:
    """Normalize a value for fuzzy matching: strip non-digits if numeric-ish, lowercase."""
    import re

    stripped = re.sub(r"[^\d]", "", val)
    if len(stripped) >= 6:  # likely a phone/ID number
        return stripped
    return val.lower().strip()


def _load_df(path: str, ext: str, sheet: str | None = None) -> tuple[Any, str | None, list[str] | None]:
    """Load DataFrame with LRU cache. Returns (df, sheet_name, all_sheet_names)."""
    import pandas as pd

    mtime = os.path.getmtime(path)
    cache_key = f"{path}:{mtime}:{sheet or '_default_'}"

    # Check cache
    with _df_cache_lock:
        if cache_key in _df_cache:
            entry = _df_cache[cache_key]
            entry["atime"] = time.time()
            return entry["df"], entry["sheet_name"], entry["all_sheets"]

    # Load fresh (outside lock — IO can be slow)
    all_sheets = None
    sheet_name = sheet

    if ext == ".csv":
        df = pd.read_csv(path)
        all_sheets = ["Sheet1"]
        sheet_name = "Sheet1"
    elif ext in (".xlsx", ".xlsm", ".xls"):
        # Get all sheet names first
        xls = pd.ExcelFile(path)
        all_sheets = xls.sheet_names
        if sheet and sheet in all_sheets:
            df = pd.read_excel(xls, sheet_name=sheet)
            sheet_name = sheet
        else:
            df = pd.read_excel(xls, sheet_name=0)
            sheet_name = all_sheets[0]
        xls.close()
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {ext}. Use CSV or Excel.")

    # Evict oldest if cache full
    with _df_cache_lock:
        if len(_df_cache) >= _DF_CACHE_MAX:
            oldest_key = min(_df_cache, key=lambda k: _df_cache[k]["atime"])
            del _df_cache[oldest_key]

        _df_cache[cache_key] = {"df": df, "sheet_name": sheet_name, "all_sheets": all_sheets, "atime": time.time()}
    return df, sheet_name, all_sheets


def _get_schema(df: Any) -> dict:
    """Return column types + sample values for Gemini guidance."""
    schema = {}
    for col in df.columns:
        dtype = str(df[col].dtype)
        sample = None
        non_null = df[col].dropna()
        if len(non_null) > 0:
            sample = str(non_null.iloc[0])
            if len(sample) > 50:
                sample = sample[:50] + "..."
        schema[col] = {"type": dtype, "sample": sample}
    return schema


class AnalyzeDataRequest(BaseModel):
    path: str
    operation: str = "describe"
    columns: str | None = None
    query: str | None = None
    sheet: str | None = None
    head: int = 20


@router.post("/analyze-data")
def analyze_data_endpoint(req: AnalyzeDataRequest) -> dict:
    """Run pandas analysis on CSV or Excel files. Cached, multi-sheet, with search."""
    import pandas as pd

    path = os.path.abspath(req.path)
    _check_safe(path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    ext = os.path.splitext(path)[1].lower()
    try:
        df, sheet_name, all_sheets = _load_df(path, ext, req.sheet)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read file: {e}")

    result = {
        "path": path,
        "shape": list(df.shape),
        "columns": list(df.columns),
        "sheet": sheet_name,
        "all_sheets": all_sheets,
    }

    op = req.operation.lower()

    if op == "describe":
        result["data"] = df.describe(include="all").to_string()
        result["schema"] = _get_schema(df)
    elif op == "head":
        result["data"] = df.head(req.head).to_string()
    elif op == "columns":
        result["data"] = _get_schema(df)
    elif op == "search":
        # Vectorized search across cells with normalization (fast even on 100MB files)
        if not req.query:
            raise HTTPException(status_code=400, detail="query required for search")
        needle = _normalize_for_search(req.query)
        matches = []

        # Determine sheets to search
        sheets_to_search = []
        if ext in (".xlsx", ".xlsm", ".xls") and not req.sheet:
            # Search ALL sheets — load each via cache
            for sn in all_sheets:
                try:
                    sdf, _, _ = _load_df(path, ext, sn)
                    sheets_to_search.append((sn, sdf))
                except Exception:
                    continue
        else:
            sheets_to_search = [(sheet_name, df)]

        for sn, sdf in sheets_to_search:
            if len(matches) >= req.head:
                break
            # Reset index for positional access (avoids slow get_loc on non-default indices)
            sdf_reset = sdf.reset_index(drop=True)
            for col in sdf_reset.columns:
                if len(matches) >= req.head:
                    break
                col_str = sdf_reset[col].astype(str).fillna("")
                if needle.isdigit() and len(needle) >= 6:
                    col_normalized = col_str.str.replace(r"[^\d]", "", regex=True)
                else:
                    col_normalized = col_str.str.lower()
                hit_mask = col_normalized.str.contains(needle, na=False, regex=False)
                hit_positions = hit_mask[hit_mask].index.tolist()  # Already positional after reset_index
                for row_pos in hit_positions:
                    if len(matches) >= req.head:
                        break
                    start = max(0, row_pos - 2)
                    end = min(len(sdf_reset), row_pos + 3)
                    # Build context as dict list (cheaper than to_string per hit)
                    context = sdf_reset.iloc[start:end].to_dict(orient="records")
                    matches.append(
                        {
                            "sheet": sn,
                            "row": row_pos + 1,
                            "column": col,
                            "value": str(col_str.iloc[row_pos]),
                            "context": context,
                        }
                    )

        result["data"] = matches
        result["matched"] = len(matches)
        if not matches:
            result["_hint"] = (
                f"No matches for '{req.query}' across {len(sheets_to_search)} sheet(s). Try different keywords or check the sheet names."
            )
    elif op == "value_counts":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for value_counts")
        col = req.columns.strip()
        if col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{col}' not found. Available: {list(df.columns)}")
        result["data"] = df[col].value_counts().head(req.head).to_string()
    elif op == "groupby":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for groupby (format: 'group_col:agg_col')")
        parts = req.columns.split(":")
        group_col = parts[0].strip()
        agg_col = parts[1].strip() if len(parts) > 1 else None
        if group_col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{group_col}' not found")
        if agg_col and agg_col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{agg_col}' not found")
        if agg_col:
            result["data"] = df.groupby(group_col)[agg_col].agg(["count", "sum", "mean", "min", "max"]).to_string()
        else:
            result["data"] = df.groupby(group_col).size().to_string()
    elif op == "filter":
        if not req.query:
            raise HTTPException(status_code=400, detail="query required for filter (e.g. 'amount > 1000')")
        try:
            filtered = df.query(req.query)
            n_matched = len(filtered)
            result["data"] = filtered.head(req.head).to_string()
            result["matched_rows"] = n_matched
            # Auto-context: show surrounding rows for sparse matches
            if 0 < n_matched <= 5:
                context_indices = set()
                for idx in filtered.index[:5]:
                    pos = df.index.get_loc(idx)
                    for offset in range(-2, 3):
                        if 0 <= pos + offset < len(df):
                            context_indices.add(df.index[pos + offset])
                result["context"] = df.loc[sorted(context_indices)].to_string()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid query: {e}")
    elif op == "corr":
        numeric_df = df.select_dtypes(include="number")
        result["data"] = numeric_df.corr().to_string()
    elif op == "sort":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for sort")
        col = req.columns.strip()
        ascending = not col.startswith("-")
        col = col.lstrip("-")
        if col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{col}' not found")
        result["data"] = df.sort_values(col, ascending=ascending).head(req.head).to_string()
    elif op == "unique":
        if not req.columns:
            raise HTTPException(status_code=400, detail="columns required for unique")
        col = req.columns.strip()
        if col not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{col}' not found")
        uniq = df[col].dropna().unique().tolist()
        result["data"] = uniq[:100]
        result["total_unique"] = len(uniq)
    elif op == "eval":
        if not req.query:
            raise HTTPException(status_code=400, detail="query required for eval (e.g. '(Qty * Price).sum()')")
        # Security: block dangerous attribute access patterns
        _blocked = ("__", "import", "exec", "eval", "compile", "globals", "locals",
                    "getattr", "setattr", "delattr", "open", "os.", "sys.", "subprocess",
                    "to_csv", "to_excel", "to_json", "to_html", "to_parquet", "to_pickle",
                    "to_sql", "to_feather", "to_clipboard", "to_latex", "to_xml",
                    "pipe", "applymap", "style", "plot", "lambda", "def ")
        query_lower = req.query.lower()
        for b in _blocked:
            if b in query_lower:
                raise HTTPException(status_code=400, detail=f"Blocked expression: '{b}' not allowed in eval queries")
        try:
            # Use pd.eval for simple expressions (arithmetic, comparisons)
            eval_result = pd.eval(req.query, local_dict={"df": df})
            if hasattr(eval_result, "to_string"):
                result["data"] = (
                    eval_result.head(req.head).to_string() if hasattr(eval_result, "head") else eval_result.to_string()
                )
            else:
                result["data"] = str(eval_result)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Eval error: {e}")
    elif op == "shape":
        result["data"] = f"{df.shape[0]} rows x {df.shape[1]} columns"
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown operation: {op}. Use: describe, head, columns, value_counts, groupby, filter, corr, sort, unique, eval, search, shape",
        )

    # Sufficiency hints per operation
    if op == "columns":
        result["_hint"] = (
            "Column info loaded. Now call the specific operation you need (filter, groupby, search, sort, etc.)."
        )
    elif op in ("filter", "search", "groupby", "sort", "value_counts", "head", "describe", "eval", "unique", "corr"):
        if not result.get("_hint"):
            result["_hint"] = "Data retrieved. Answer the user's question with these results."

    return result


# --- Table Extraction from PDFs (pdfplumber) ---


@router.post("/extract-tables")
def extract_tables_endpoint(
    path: str = Query(..., description="Path to PDF file"),
    pages: str | None = Query(None, description="Page range: '1-5', '3', 'all'. Default: all"),
) -> dict:
    """Extract structured tables from a PDF using pdfplumber."""
    path = os.path.abspath(path)
    _check_safe(path)
    if not os.path.exists(path):
        raise HTTPException(404, f"File not found: {path}")
    if not path.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files supported")

    import pdfplumber

    try:
        pdf = pdfplumber.open(path)
    except Exception as e:
        return {"error": f"Cannot open PDF: {e}"}

    with pdf:
        total_pages = len(pdf.pages)

        # Parse page range
        page_indices = []
        if not pages or pages == "all":
            page_indices = list(range(total_pages))
        elif "-" in pages:
            parts = pages.split("-")
            start = max(int(parts[0]) - 1, 0)
            end = min(int(parts[1]), total_pages)
            page_indices = list(range(start, end))
        else:
            p = int(pages) - 1
            if 0 <= p < total_pages:
                page_indices = [p]

        tables = []
        for pi in page_indices:
            page = pdf.pages[pi]
            page_tables = page.extract_tables()
            for ti, table in enumerate(page_tables):
                if not table or len(table) < 2:
                    continue
                # First row as headers, rest as data
                headers = [str(c).strip() if c else "" for c in table[0]]
                rows = []
                for row in table[1:]:
                    rows.append([str(c).strip() if c else "" for c in row])
                tables.append(
                    {
                        "page": pi + 1,
                        "table_index": ti + 1,
                        "headers": headers,
                        "rows": rows,
                        "row_count": len(rows),
                    }
                )

    if not tables:
        return {
            "path": path,
            "total_pages": total_pages,
            "tables": [],
            "error": "No tables found in this PDF. Tables must be structured (not scanned images).",
            "_hint": "No tables detected. If the PDF is scanned, try OCR first with ocr(path).",
        }

    return {
        "path": path,
        "total_pages": total_pages,
        "tables_found": len(tables),
        "tables": tables,
        "_hint": "Tables extracted. Present them clearly or use analyze_data for further analysis.",
    }


@router.post("/pdf-to-excel")
def pdf_to_excel_endpoint(
    path: str = Query(..., description="Path to PDF file"),
    output_path: str | None = Query(None, description="Output .xlsx path. Default: same name with .xlsx"),
    pages: str | None = Query(None, description="Page range: '1-5', '3', 'all'. Default: all"),
) -> dict:
    """Extract tables from a PDF and save as an Excel (.xlsx) file. Each table becomes a sheet."""
    import pdfplumber

    path = os.path.abspath(path)
    _check_safe(path)
    if not os.path.exists(path):
        raise HTTPException(404, f"File not found: {path}")
    if not path.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files supported")

    output = os.path.abspath(output_path) if output_path else path.rsplit(".", 1)[0] + ".xlsx"
    _check_safe(output)
    os.makedirs(os.path.dirname(output), exist_ok=True)

    try:
        pdf = pdfplumber.open(path)
    except Exception as e:
        return {"error": f"Cannot open PDF: {e}"}

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    table_count = 0

    with pdf:
        total_pages = len(pdf.pages)

        # Parse page range
        page_indices = []
        if not pages or pages == "all":
            page_indices = list(range(total_pages))
        elif "-" in pages:
            parts = pages.split("-")
            start = max(int(parts[0]) - 1, 0)
            end = min(int(parts[1]), total_pages)
            page_indices = list(range(start, end))
        else:
            p = int(pages) - 1
            if 0 <= p < total_pages:
                page_indices = [p]

        for pi in page_indices:
            page = pdf.pages[pi]
            page_tables = page.extract_tables()
            for ti, table in enumerate(page_tables):
                if not table or len(table) < 2:
                    continue
                table_count += 1
                sheet_name = f"P{pi + 1}_T{ti + 1}"[:31]  # Excel sheet name max 31 chars
                ws = wb.create_sheet(title=sheet_name)
                for row in table:
                    ws.append([str(c).strip() if c else "" for c in row])

    if table_count == 0:
        return {
            "path": path,
            "error": "No tables found in this PDF.",
            "_hint": "No tables detected. If scanned, try OCR first.",
        }

    wb.save(output)
    record_generated_file(output, "pdf_to_excel", f"Excel from PDF: {table_count} tables")
    return {
        "success": True,
        "path": output,
        "tables_exported": table_count,
        "total_pages": total_pages,
        "_hint": f"{table_count} table(s) exported to Excel. Send the file or use analyze_data to work with it.",
    }
