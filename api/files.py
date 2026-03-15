"""File operation endpoints: list, info, read, move, batch move, create, delete, grep, duplicates, batch rename, path registry."""

from __future__ import annotations

import logging
import os
import pathlib
import shutil
import sqlite3
import threading
from datetime import datetime

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel

from api.helpers import (
    EXCEL_EXTS,
    IMAGE_EXTS,
    MAX_READ_SIZE,
    MAX_TEXT_CHARS,
    OFFICE_EXTS,
    PDF_EXTS,
    TEXT_EXTS,
    _background_index,
    _check_safe,
    _get_conn,
    _human_date,
    _human_size,
)
from indexing_service import index_single_file
from job_service import (
    cancel_jobs_for_target,
    create_job,
    get_or_create_job,
    is_job_cancelling,
    list_jobs,
    mark_job_cancelled,
    mark_job_completed,
    mark_job_failed,
    mark_job_running,
    request_job_cancel,
    update_job_progress,
    update_job_stage,
)

router = APIRouter()
logger = logging.getLogger(__name__)


def _update_path_references(conn, old_path: str, new_path: str) -> None:
    """Propagate a file path change across all path-keyed tables."""
    conn.execute("UPDATE documents SET path = ? WHERE path = ?", (new_path, old_path))
    for stmt in [
        ("UPDATE video_embeddings SET video_path = ? WHERE video_path = ?", (new_path, old_path)),
        ("UPDATE photo_classifications SET path = ? WHERE path = ?", (new_path, old_path)),
        ("UPDATE photo_scores SET path = ? WHERE path = ?", (new_path, old_path)),
        ("UPDATE image_embeddings SET path = ? WHERE path = ?", (new_path, old_path)),
        ("UPDATE face_cache SET image_path = ? WHERE image_path = ?", (new_path, old_path)),
        ("UPDATE file_paths SET path = ? WHERE path = ?", (new_path, old_path)),
        ("UPDATE generated_files SET path = ? WHERE path = ?", (new_path, old_path)),
        ("UPDATE known_faces SET source_image = ? WHERE source_image = ?", (new_path, old_path)),
    ]:
        try:
            conn.execute(*stmt)
        except sqlite3.Error:
            continue


# --- List files (enhanced with sort/filter) ---

# Extension groups for filter_type
_TYPE_EXTENSIONS = {
    "image": {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif", ".tiff", ".tif", ".svg", ".ico", ".heic"},
    "document": {".pdf", ".docx", ".doc", ".txt", ".epub", ".rtf", ".odt"},
    "spreadsheet": {".xlsx", ".xls", ".csv", ".ods"},
    "presentation": {".pptx", ".ppt", ".odp"},
    "video": {".mp4", ".mkv", ".avi", ".mov", ".wmv", ".flv", ".webm"},
    "audio": {".mp3", ".wav", ".flac", ".aac", ".ogg", ".wma", ".m4a"},
    "archive": {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2"},
}


@router.get("/list_files")
def list_files_endpoint(
    folder: str = Query(..., description="Folder path"),
    sort_by: str = Query("name", description="Sort by: name, date, size"),
    filter_ext: str | None = Query(None, description="Filter by extension: .pdf, .xlsx"),
    filter_type: str | None = Query(
        None, description="Filter by type: image, document, spreadsheet, presentation, video, audio, archive"
    ),
    name_contains: str | None = Query(None, description="Filter by filename containing this text (case-insensitive)"),
    recursive: bool = Query(False, description="Search subdirectories recursively"),
    limit: int = Query(200, ge=1, le=2000, description="Max entries"),
) -> dict:
    """List contents of a folder with sorting and filtering."""
    folder = os.path.abspath(folder)
    _check_safe(folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    # Resolve filter_type to a set of extensions
    allowed_exts = None
    if filter_type and filter_type.lower() in _TYPE_EXTENSIONS:
        allowed_exts = _TYPE_EXTENSIONS[filter_type.lower()]
    elif filter_ext:
        allowed_exts = {filter_ext.lower()}

    name_lower = name_contains.lower() if name_contains else None
    entries = []
    _SCAN_CAP = 10000  # cap total files scanned to prevent multi-minute walks

    def _process_entry(de_name: str, de_path: str, is_scan_entry: bool = True) -> None:
        """Process a directory entry and add to entries if it matches filters."""
        if name_lower and name_lower not in de_name.lower():
            return
        if allowed_exts:
            ext = os.path.splitext(de_name)[1].lower()
            if ext not in allowed_exts:
                return
        try:
            st = os.stat(de_path)
            is_dir = os.path.isdir(de_path)
            if allowed_exts and is_dir:
                return
            entries.append(
                {
                    "name": de_name,
                    "path": de_path,
                    "is_dir": is_dir,
                    "size": st.st_size if not is_dir else 0,
                    "size_human": _human_size(st.st_size) if not is_dir else "-",
                    "modified": _human_date(st.st_mtime),
                    "modified_ts": st.st_mtime,
                }
            )
        except OSError:
            pass

    if recursive and name_lower:
        # Use native OS search — fast, no cap, searches everything
        import re
        import subprocess

        _found_via_native = False
        try:
            # Sanitize name_contains — strip shell metacharacters
            _safe_name = re.sub(r'[&|;$`"\'\\<>()!^%*?\[\]\r\n\t]', '', name_contains)

            # Use find command (safe: list args, no shell) for both WSL and Linux paths
            cmd = ["find", folder, "-maxdepth", "5", "-iname", f"*{_safe_name}*", "-not", "-path", "*/.*"]
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            for line in r.stdout.strip().split("\n"):
                if line:
                    _process_entry(os.path.basename(line), line)
                    if len(entries) >= limit:
                        break
            _found_via_native = True
        except Exception:
            pass
        if not _found_via_native:
            # Fallback to os.walk
            scanned = 0
            for root, dirs, files in os.walk(folder):
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                for fname in files:
                    if scanned >= _SCAN_CAP:
                        break
                    _process_entry(fname, os.path.join(root, fname))
                    scanned += 1
                if scanned >= _SCAN_CAP:
                    break
    elif recursive:
        scanned = 0
        for root, dirs, files in os.walk(folder):
            dirs[:] = [d for d in dirs if not d.startswith(".")]
            for fname in files:
                if scanned >= _SCAN_CAP:
                    break
                _process_entry(fname, os.path.join(root, fname))
                scanned += 1
            if scanned >= _SCAN_CAP:
                break
    else:
        # Use os.scandir for performance (caches stat results)
        try:
            with os.scandir(folder) as scanner:
                for de in scanner:
                    _process_entry(de.name, de.path)
        except OSError:
            import subprocess

            try:
                result = subprocess.run(["ls", "-1", folder], capture_output=True, text=True, timeout=10)
                for name in result.stdout.strip().split("\n"):
                    if name:
                        _process_entry(name, os.path.join(folder, name))
            except Exception:
                pass

    # Sort
    if sort_by == "date":
        entries.sort(key=lambda e: e["modified_ts"], reverse=True)
    elif sort_by == "size":
        entries.sort(key=lambda e: e["size"], reverse=True)
    else:
        entries.sort(key=lambda e: (not e["is_dir"], e["name"].lower()))

    total = len(entries)
    entries = entries[:limit]

    result = {"folder": folder, "total": total, "showing": len(entries), "entries": entries}
    # Add summary stats so Gemini can answer without parsing all entries
    if entries and sort_by == "size" and not entries[0].get("is_dir"):
        result["largest"] = f"{entries[0]['name']} ({entries[0]['size_human']})"
    if total > len(entries):
        result["_hint"] = (
            f"{total} items found, showing {len(entries)}. Use name_contains or filter_ext to narrow down."
        )
    elif total > 0:
        result["_hint"] = f"{total} item(s) listed. Use this to answer or plan next action."
    else:
        result["_hint"] = "Folder is empty. Check if the path is correct."
    return result


# --- Grep (content search inside files) ---


class GrepRequest(BaseModel):
    pattern: str
    folder: str
    file_filter: str | None = None  # e.g. "*.txt", "*.xlsx"
    max_results: int = 20


@router.post("/grep")
def grep_endpoint(req: GrepRequest) -> dict:
    """Search inside files by text pattern using OS grep."""
    import subprocess

    folder = os.path.abspath(req.folder)
    _check_safe(folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    cmd = ["grep", "-ril", "--include=" + (req.file_filter or "*"), req.pattern, folder]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        files = [f for f in r.stdout.strip().split("\n") if f][: req.max_results]
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=408, detail="Search timed out (folder too large)")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not files:
        return {"pattern": req.pattern, "folder": folder, "matches": 0, "files": []}

    # Get matching lines for context
    results = []
    for fpath in files:
        try:
            r2 = subprocess.run(
                ["grep", "-in", "--max-count=3", req.pattern, fpath], capture_output=True, text=True, timeout=5
            )
            lines = [l.strip() for l in r2.stdout.strip().split("\n") if l.strip()][:3]
            results.append({"path": fpath, "name": os.path.basename(fpath), "matches": lines})
        except Exception:
            results.append({"path": fpath, "name": os.path.basename(fpath), "matches": []})

    return {"pattern": req.pattern, "folder": folder, "matches": len(results), "files": results}


# --- File info ---


@router.get("/file_info")
def file_info_endpoint(
    path: str = Query(..., description="File or folder path"),
) -> dict:
    """Get detailed file/folder metadata and indexed status."""
    path = os.path.abspath(path)
    _check_safe(path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Not found: {path}")

    stat = os.stat(path)
    is_dir = os.path.isdir(path)

    info = {
        "name": os.path.basename(path),
        "path": path,
        "is_dir": is_dir,
        "size": stat.st_size,
        "size_human": _human_size(stat.st_size),
        "created": _human_date(stat.st_ctime),
        "modified": _human_date(stat.st_mtime),
        "extension": os.path.splitext(path)[1].lower() if not is_dir else None,
    }

    if is_dir:
        # Count files in directory (single-pass with scandir)
        try:
            file_count = folder_count = 0
            with os.scandir(path) as it:
                for entry in it:
                    if entry.is_file():
                        file_count += 1
                    elif entry.is_dir():
                        folder_count += 1
            info["file_count"] = file_count
            info["folder_count"] = folder_count
        except OSError:
            info["file_count"] = 0
            info["folder_count"] = 0
    else:
        # Check if indexed in database
        conn = _get_conn()
        row = conn.execute("SELECT id, file_type, active FROM documents WHERE path = ?", (path,)).fetchone()
        if row:
            info["indexed"] = True
            info["document_id"] = row["id"]
            info["file_type"] = row["file_type"]
            info["active"] = bool(row["active"])
        else:
            info["indexed"] = False
            info["_hint"] = "File not indexed. Use index_file to make it searchable."

    return info


# --- Read file from disk (images -> base64, docs -> text) ---


class ReadFileRequest(BaseModel):
    path: str


@router.post("/read_file")
def read_file_endpoint(req: ReadFileRequest) -> dict:
    """Read actual file from disk. Images return base64, documents return text."""
    import base64

    path = os.path.abspath(req.path)
    _check_safe(path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    if os.path.isdir(path):
        raise HTTPException(status_code=400, detail="Cannot read a directory")

    file_size = os.path.getsize(path)
    if file_size > MAX_READ_SIZE:
        raise HTTPException(
            status_code=400, detail=f"File too large ({_human_size(file_size)}). Max {_human_size(MAX_READ_SIZE)}"
        )

    ext = os.path.splitext(path)[1].lower()

    # Fire-and-forget: auto-index text-extractable files for future semantic search
    if ext in TEXT_EXTS | PDF_EXTS | OFFICE_EXTS | EXCEL_EXTS:
        _background_index(path)

    # Images -> base64 for Gemini vision + cached caption if indexed
    if ext in IMAGE_EXTS:
        with open(path, "rb") as f:
            data = base64.b64encode(f.read()).decode("ascii")
        mime_map = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".bmp": "image/bmp",
            ".webp": "image/webp",
            ".gif": "image/gif",
            ".tiff": "image/tiff",
            ".tif": "image/tiff",
            ".heic": "image/heic",
        }
        result = {
            "type": "image",
            "mime_type": mime_map.get(ext, "image/jpeg"),
            "data": data,
            "path": path,
            "size": file_size,
        }
        # DB-first: enrich with cached caption if already indexed (free)
        try:
            conn = _get_conn()
            doc = conn.execute("SELECT hash FROM documents WHERE path = ? AND active = 1", (path,)).fetchone()
            if doc:
                cr = conn.execute("SELECT text FROM content WHERE hash = ?", (doc["hash"],)).fetchone()
                if cr and cr["text"]:
                    result["caption"] = cr["text"]
                    result["_hint"] = "Image has cached caption from index — no need to re-caption."
        except Exception:
            pass
        return result

    # Text files -> raw content
    if ext in TEXT_EXTS:
        try:
            with open(path, encoding="utf-8", errors="replace") as f:
                content = f.read(MAX_TEXT_CHARS + 100)
            truncated = len(content) > MAX_TEXT_CHARS
            if truncated:
                content = content[:MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
            resp = {"type": "text", "content": content, "path": path, "truncated": truncated}
            if truncated:
                resp["_hint"] = "Content truncated. Use index_file + search_documents to find specific sections."
            return resp
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read file: {e}")

    # PDFs -> extract with PyMuPDF, fallback to OCR for scanned PDFs
    if ext in PDF_EXTS:
        try:
            import pymupdf4llm

            content = pymupdf4llm.to_markdown(path)
            # If empty text -> likely scanned PDF, try OCR
            if not content or not content.strip():
                ocr_result = _ocr_single(path)
                if ocr_result and ocr_result.get("text"):
                    content = ocr_result["text"]
            truncated = len(content) > MAX_TEXT_CHARS
            if truncated:
                content = content[:MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
            resp = {"type": "text", "content": content, "path": path, "truncated": truncated}
            if truncated:
                resp["_hint"] = "Content truncated. Use index_file + search_documents to find specific sections."
            return resp
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read PDF: {e}")

    # Office docs -> MarkItDown
    if ext in OFFICE_EXTS:
        try:
            from markitdown import MarkItDown

            md = MarkItDown(enable_plugins=False)
            result = md.convert(path)
            content = result.text_content
            truncated = len(content) > MAX_TEXT_CHARS
            if truncated:
                content = content[:MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
            resp = {"type": "text", "content": content, "path": path, "truncated": truncated}
            if truncated:
                resp["_hint"] = "Content truncated. Use index_file + search_documents to find specific sections."
            return resp
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read document: {e}")

    # Excel -> all-sheets summary (sheet names + row counts + first 5 rows each)
    if ext in EXCEL_EXTS:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_count = ws.max_row or 0
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 6:
                        break
                    rows.append([str(v) if v is not None else "" for v in row])
                lines = [f"## Sheet: {sheet_name} ({row_count} rows)"]
                if rows:
                    lines.append("| " + " | ".join(rows[0]) + " |")
                    lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
                    for row in rows[1:]:
                        padded = row + [""] * (len(rows[0]) - len(row))
                        lines.append("| " + " | ".join(padded[: len(rows[0])]) + " |")
                    if row_count > 6:
                        lines.append(f"... ({row_count - 6} more rows)")
                parts.append("\n".join(lines))
            wb.close()
            content = "\n\n".join(parts)
            return {
                "type": "text",
                "content": content,
                "path": path,
                "truncated": False,
                "_hint": "For data analysis (filter, groupby, search values), use analyze_data instead.",
            }
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read Excel: {e}")

    # Unknown type -> try as text
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            content = f.read(MAX_TEXT_CHARS + 100)
        truncated = len(content) > MAX_TEXT_CHARS
        if truncated:
            content = content[:MAX_TEXT_CHARS] + "\n\n[... truncated ...]"
        return {"type": "text", "content": content, "path": path, "truncated": truncated}
    except Exception:
        raise HTTPException(status_code=400, detail=f"Cannot read file type: {ext}")


# --- OCR helper (used by read_file for scanned PDFs, and by /ocr in media.py) ---


def _ocr_single(path: str) -> dict:
    """OCR a single file (image or PDF). Checks DB first — if already indexed, returns cached text."""
    path = os.path.abspath(path)
    if not os.path.exists(path):
        return {"error": f"File not found: {path}"}

    # DB-first: check if already indexed (free — skip expensive OCR)
    try:
        conn = _get_conn()
        doc = conn.execute("SELECT hash FROM documents WHERE path = ? AND active = 1", (path,)).fetchone()
        if doc:
            content_row = conn.execute("SELECT text FROM content WHERE hash = ?", (doc["hash"],)).fetchone()
            if content_row and content_row["text"] and len(content_row["text"].strip()) > 10:
                return {
                    "path": path,
                    "text": content_row["text"],
                    "method": "cached_index",
                    "_hint": "Text from DB index (no re-processing needed).",
                }
    except Exception:
        pass

    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        from extractors import extract_pdf

        result = extract_pdf(path)
        if result is None:
            return {"error": "Failed to extract text from PDF"}
        return {"path": path, "text": result["text"], "pages": result.get("page_count", 0), "method": "pdf_ocr"}
    elif ext in IMAGE_EXTS:
        try:
            from PIL import Image

            from extractors import _HAS_TESSERACT, _get_gemini, _ocr_gemini, _ocr_tesseract, _preprocess_image

            img_orig = Image.open(path).convert("RGB")
            try:
                img = _preprocess_image(img_orig)
                _gemini = _get_gemini()
                text = _ocr_gemini([img]) if _gemini else (_ocr_tesseract([img]) if _HAS_TESSERACT else "")
                method = "gemini_ocr" if _gemini else ("tesseract_ocr" if _HAS_TESSERACT else "none")
                return {
                    "path": path,
                    "text": text,
                    "method": method,
                    "_hint": "Use index_file to make this text searchable, or search_documents if already indexed.",
                }
            finally:
                img_orig.close()
                if img is not img_orig:
                    img.close()
        except Exception as e:
            return {"error": f"OCR failed: {e}"}
    else:
        return {"error": f"OCR not supported for {ext}. Use images or PDFs."}


# --- Move file ---


class MoveFileRequest(BaseModel):
    source: str
    destination: str
    is_copy: bool = False


@router.post("/move_file")
def move_file_endpoint(req: MoveFileRequest) -> dict:
    """Move, copy, or rename a file."""
    src = os.path.abspath(req.source)
    _check_safe(src)
    dest = os.path.abspath(req.destination)
    _check_safe(dest)

    if not os.path.exists(src):
        raise HTTPException(status_code=404, detail=f"Source not found: {src}")

    # If dest is a directory, move file INTO it
    if os.path.isdir(dest):
        dest = os.path.join(dest, os.path.basename(src))

    # Don't overwrite
    if os.path.exists(dest) and os.path.abspath(src) != os.path.abspath(dest):
        raise HTTPException(status_code=409, detail=f"Destination already exists: {dest}")

    # Create parent directory if needed
    os.makedirs(os.path.dirname(dest), exist_ok=True)

    if req.is_copy:
        if os.path.isdir(src):
            shutil.copytree(src, dest)
        else:
            shutil.copy2(src, dest)
        action = "copied"
    else:
        shutil.move(src, dest)
        action = "moved"

        # Update database paths if the file was indexed
        conn = _get_conn()
        _update_path_references(conn, src, dest)
        conn.commit()

    return {
        "success": True,
        "source": src,
        "destination": dest,
        "action": action,
        "_hint": f"{action.capitalize()} {os.path.basename(src)} → {dest}. Report this to user.",
    }


# --- Batch move/copy files ---


class BatchMoveRequest(BaseModel):
    sources: list[str]
    destination: str
    is_copy: bool = False


@router.post("/batch_move")
def batch_move_endpoint(req: BatchMoveRequest) -> dict:
    """Move or copy multiple files to a destination folder."""
    dest_folder = os.path.abspath(req.destination)
    _check_safe(dest_folder)
    os.makedirs(dest_folder, exist_ok=True)

    results = {"moved": [], "skipped": [], "errors": []}
    conn = _get_conn()

    db_updates = []  # collect (dest, src) pairs for single commit

    for src_path in req.sources:
        src = os.path.abspath(src_path)
        _check_safe(src)
        dest = os.path.join(dest_folder, os.path.basename(src))
        try:
            if not os.path.exists(src):
                results["skipped"].append(os.path.basename(src))
                continue
            if os.path.exists(dest):
                results["skipped"].append(os.path.basename(src))
                continue
            if req.is_copy:
                shutil.copy2(src, dest)
            else:
                shutil.move(src, dest)
                db_updates.append((dest, src))
            results["moved"].append(os.path.basename(src))
        except Exception as e:
            results["errors"].append(f"{os.path.basename(src)}: {e}")

    # Single DB commit after all filesystem moves succeed
    if db_updates and not req.is_copy:
        try:
            for dest, src in db_updates:
                _update_path_references(conn, src, dest)
            conn.commit()
        except Exception:
            conn.rollback()
    action = "copied" if req.is_copy else "moved"
    moved_count = len(results["moved"])
    skipped_count = len(results["skipped"])
    error_count = len(results["errors"])
    # Honest result hint — forces truthful reporting (OpenClaw pattern: make lies impossible)
    if moved_count == 0:
        hint = f"WARNING: 0 files were {action}. {skipped_count} skipped (already exist at destination or source missing). {error_count} errors. Tell the user NO files were {action}."
    else:
        hint = f"ACTUALLY {action} {moved_count} files to {dest_folder}. {skipped_count} skipped, {error_count} errors. Report these exact numbers."
    return {
        "success": moved_count > 0,
        "action": action,
        "destination": dest_folder,
        "moved_count": moved_count,
        "skipped_count": skipped_count,
        "error_count": error_count,
        "errors": results["errors"][:5],
        "_hint": hint,
    }


# --- Create folder ---


class CreateFolderRequest(BaseModel):
    path: str


@router.post("/create_folder")
def create_folder_endpoint(req: CreateFolderRequest) -> dict:
    """Create a new directory (with parents if needed)."""
    path = os.path.abspath(req.path)
    _check_safe(path)

    already_existed = os.path.exists(path)
    os.makedirs(path, exist_ok=True)

    return {"success": True, "path": path, "already_existed": already_existed}


# --- Delete file ---


class DeleteFileRequest(BaseModel):
    path: str


@router.post("/delete_file")
def delete_file_endpoint(req: DeleteFileRequest) -> dict:
    """Delete a file (not folders — too dangerous)."""
    path = os.path.abspath(req.path)
    _check_safe(path)

    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Not found: {path}")
    if os.path.isdir(path):
        raise HTTPException(status_code=400, detail="Cannot delete directories (safety). Only files.")

    os.remove(path)

    # Soft-delete from database if indexed, clean up related tables
    conn = _get_conn()
    doc_row = conn.execute("SELECT id FROM documents WHERE path = ?", (path,)).fetchone()
    conn.execute("UPDATE documents SET active = 0 WHERE path = ?", (path,))
    if doc_row:
        doc_id = doc_row["id"]
        conn.execute("DELETE FROM facts WHERE document_id = ?", (doc_id,))
        conn.execute("DELETE FROM chunks WHERE document_id = ?", (doc_id,))
    # Clean up path-based tables
    for stmt in [
        ("DELETE FROM image_embeddings WHERE path = ?", (path,)),
        ("DELETE FROM face_cache WHERE image_path = ?", (path,)),
        ("DELETE FROM file_paths WHERE path = ?", (path,)),
        ("DELETE FROM photo_scores WHERE path = ?", (path,)),
        ("DELETE FROM photo_classifications WHERE path = ?", (path,)),
    ]:
        try:
            conn.execute(*stmt)
        except Exception:
            pass
    conn.commit()

    return {"success": True, "path": path}


# --- Find duplicates ---


class FindDuplicatesRequest(BaseModel):
    folder: str


@router.post("/find-duplicates")
def find_duplicates_endpoint(req: FindDuplicatesRequest) -> dict:
    """Find duplicate files in a folder by content hash."""
    import hashlib

    folder = os.path.abspath(req.folder)
    _check_safe(folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    hashes = {}  # hash -> [paths]
    scanned = 0
    _MAX_FILES_SCAN = 20000  # cap to prevent multi-minute scans on huge folder trees
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for name in files:
            if scanned >= _MAX_FILES_SCAN:
                break
            fp = os.path.join(root, name)
            try:
                size = os.path.getsize(fp)
                if size == 0:
                    continue
                # Hash first 8KB + size for speed
                with open(fp, "rb") as f:
                    data = f.read(8192)
                key = f"{size}:{hashlib.md5(data).hexdigest()}"
                hashes.setdefault(key, []).append(fp)
                scanned += 1
            except OSError:
                continue
        if scanned >= _MAX_FILES_SCAN:
            break

    duplicates = {k: v for k, v in hashes.items() if len(v) > 1}
    groups = [{"files": paths, "count": len(paths)} for paths in duplicates.values()]
    groups.sort(key=lambda g: g["count"], reverse=True)

    resp = {
        "folder": folder,
        "scanned": scanned,
        "duplicate_groups": len(groups),
        "total_duplicates": sum(g["count"] - 1 for g in groups),
        "groups": groups[:50],  # cap at 50 groups
    }
    if groups:
        resp["_hint"] = "Use delete_file to remove duplicates (keep the original). Ask user which copies to remove."
    return resp


# --- Batch rename ---


class BatchRenameRequest(BaseModel):
    folder: str
    pattern: str  # regex or glob pattern to match
    replace: str  # replacement string
    dry_run: bool = True  # preview only by default


@router.post("/batch-rename")
def batch_rename_endpoint(req: BatchRenameRequest) -> dict:
    """Rename files in a folder matching a pattern. dry_run=true (default) shows preview only."""
    import re

    folder = os.path.abspath(req.folder)
    _check_safe(folder)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Not a directory: {folder}")

    renamed = []
    errors = []
    try:
        regex = re.compile(req.pattern)
    except re.error as e:
        raise HTTPException(status_code=400, detail=f"Invalid regex pattern: {e}")

    db_updates = []  # collect (new_path, old_path) for single DB commit

    for name in os.listdir(folder):
        new_name = regex.sub(req.replace, name)
        if new_name != name:
            old_path = os.path.join(folder, name)
            new_path = os.path.join(folder, new_name)
            if os.path.isfile(old_path) and not os.path.exists(new_path):
                if req.dry_run:
                    renamed.append({"old": name, "new": new_name})
                else:
                    try:
                        os.rename(old_path, new_path)
                        renamed.append({"old": name, "new": new_name})
                        db_updates.append((new_path, old_path))
                    except OSError as e:
                        errors.append({"file": name, "error": str(e)})

    # Update DB paths in single commit (same as batch_move)
    if db_updates:
        conn = _get_conn()
        try:
            for new_path, old_path in db_updates:
                _update_path_references(conn, old_path, new_path)
            conn.commit()
        except Exception:
            conn.rollback()

    resp = {
        "folder": folder,
        "renamed": len(renamed),
        "errors": len(errors),
        "details": renamed[:50],
        "dry_run": req.dry_run,
    }
    if req.dry_run and renamed:
        resp["_hint"] = f"Preview: {len(renamed)} files would be renamed. Call again with dry_run=false to execute."
    return resp


# --- Search generated files ---


@router.get("/search-generated-files")
def search_generated_files_endpoint(
    query: str = Query("", description="Search term (matches path and description)"),
    tool_name: str = Query("", description="Filter by tool name"),
    limit: int = Query(50, ge=1, le=200),
) -> dict:
    """Search files previously created by Pinpoint tools."""
    conn = _get_conn()
    conditions = []
    params: list = []

    if query:
        like = f"%{query}%"
        conditions.append("(path LIKE ? OR description LIKE ?)")
        params.extend([like, like])
    if tool_name:
        conditions.append("tool_name = ?")
        params.append(tool_name)

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    rows = conn.execute(
        f"SELECT id, path, tool_name, description, created_at FROM generated_files {where} ORDER BY created_at DESC LIMIT ?",
        params + [limit],
    ).fetchall()

    results = []
    for r in rows:
        entry = {"id": r[0], "path": r[1], "tool_name": r[2], "description": r[3], "created_at": r[4]}
        entry["exists"] = os.path.exists(r[1])
        results.append(entry)

    return {
        "results": results,
        "count": len(results),
        "_hint": f"{len(results)} generated file(s) found." if results else "No generated files match your query.",
    }


# --- File path registry (auto-scan common folders on startup) ---

_SKIP_DIRS = {
    ".git", "node_modules", "__pycache__", ".cache", ".venv", "venv",
    ".tox", ".mypy_cache", ".pytest_cache", "AppData", "$Recycle.Bin",
    "System Volume Information", ".Trash", ".local", ".config",
}

_scan_status = {"running": False, "total": 0, "folders": [], "indexed": 0}
_RESCAN_INTERVAL = 3600  # re-scan every 60 minutes
_AUTO_INDEX_EXTS = {".pdf", ".docx", ".xlsx", ".pptx", ".epub", ".txt", ".csv", ".log", ".md"}
_MAX_AUTO_INDEX_SIZE = 100 * 1024 * 1024  # 100MB — skip huge files


_SKIP_USERS = {"All Users", "Default", "Default User", "Public", "TEMP"}


def _get_common_folders() -> list[str]:
    """Detect OS-standard user folders."""
    home = pathlib.Path.home()
    candidates = ["Documents", "Desktop", "Downloads", "Pictures", "Videos", "Music"]
    folders = []
    for name in candidates:
        p = home / name
        if p.is_dir():
            folders.append(str(p))
    # On WSL/Windows, also check mounted drives for real user folders
    for drive in ["/mnt/c", "/mnt/d", "/mnt/e"]:
        users = os.path.join(drive, "Users")
        if os.path.isdir(users):
            for user_dir in os.listdir(users):
                if user_dir in _SKIP_USERS or user_dir.startswith("TEMP"):
                    continue
                for name in candidates:
                    p = os.path.join(users, user_dir, name)
                    if os.path.isdir(p) and p not in folders:
                        folders.append(p)
    return folders


def _walk_folder(folder: str):
    """Walk a folder yielding (path, filename, ext, size, mtime). Falls back to `find` on WSL I/O errors."""
    found_any = False
    try:
        for root, dirs, files in os.walk(folder):
            dirs[:] = [d for d in dirs if d not in _SKIP_DIRS and not d.startswith(".")]
            for fname in files:
                fpath = os.path.join(root, fname)
                ext = os.path.splitext(fname)[1].lower()
                try:
                    st = os.stat(fpath)
                    found_any = True
                    yield fpath, fname, ext, st.st_size, st.st_mtime
                except OSError:
                    continue
    except OSError:
        pass

    # Fallback: WSL can't os.walk some Windows folders (I/O error) but `find` works
    if not found_any:
        try:
            import subprocess

            result = subprocess.run(
                ["find", folder, "-maxdepth", "3", "-type", "f"],
                capture_output=True, text=True, timeout=60,
            )
            for line in result.stdout.strip().split("\n"):
                if not line:
                    continue
                # Skip hidden/junk dirs
                parts = line.split("/")
                if any(p in _SKIP_DIRS or p.startswith(".") for p in parts):
                    continue
                fname = os.path.basename(line)
                ext = os.path.splitext(fname)[1].lower()
                try:
                    st = os.stat(line)
                    yield line, fname, ext, st.st_size, st.st_mtime
                except OSError:
                    continue
        except Exception:
            pass


def _get_watched_folders() -> list[str]:
    """Get user-registered watched folders from DB."""
    try:
        conn = _get_conn()
        rows = conn.execute("SELECT path FROM watched_folders").fetchall()
        return [r["path"] for r in rows if os.path.isdir(r["path"])]
    except sqlite3.Error as e:
        logger.exception("watched_folders_load_failed")
        return []


def scan_paths_background():
    """Walk common folders for path registry + auto-index ONLY watched folders. Repeats every 60 min."""
    if _scan_status["running"]:
        return

    def _scan_loop():
        import time

        while True:
            _scan_status["running"] = True
            conn = _get_conn()
            scan_job_id, created = get_or_create_job(conn, "path_registry_scan", target_path="", current_stage="queued")
            if created:
                mark_job_running(conn, scan_job_id, current_stage="walking")
            else:
                update_job_stage(conn, scan_job_id, "walking")
            # Path registry: common folders (file lookup only, NO indexing)
            registry_folders = _get_common_folders()
            # Watched folders: user-chosen, these get auto-indexed + embedded
            watched = _get_watched_folders()
            watched_set = set(watched)
            all_folders = list(set(registry_folders + watched))
            update_job_progress(
                conn,
                scan_job_id,
                total_items=len(all_folders),
                completed_items=0,
                details={"watched_folders": len(watched), "paths_scanned": 0},
                current_stage="walking",
            )
            _scan_status["folders"] = all_folders
            _scan_status["total"] = 0
            _scan_status["indexed"] = 0

            try:
                now = datetime.now().isoformat()
                batch = []
                count = 0
                to_index = []  # files to auto-index (watched folders ONLY)

                for folder_index, folder in enumerate(all_folders, start=1):
                    is_watched = folder in watched_set
                    folder_files = list(_walk_folder(folder))
                    for fpath, fname, ext, size, mtime in folder_files:
                        batch.append((fpath, fname, ext, size, mtime, now))
                        # Only auto-index files in watched folders
                        if is_watched and ext in _AUTO_INDEX_EXTS and size <= _MAX_AUTO_INDEX_SIZE:
                            to_index.append(fpath)
                        if len(batch) >= 1000:
                            conn.executemany(
                                "INSERT OR REPLACE INTO file_paths (path, filename, ext, size_bytes, modified_at, scanned_at) VALUES (?, ?, ?, ?, ?, ?)",
                                batch,
                            )
                            conn.commit()
                            count += len(batch)
                            _scan_status["total"] = count
                            batch = []
                    update_job_progress(
                        conn,
                        scan_job_id,
                        total_items=len(all_folders),
                        completed_items=folder_index,
                        details={"watched_folders": len(watched), "paths_scanned": count + len(batch)},
                        current_stage="walking",
                    )
                if batch:
                    conn.executemany(
                        "INSERT OR REPLACE INTO file_paths (path, filename, ext, size_bytes, modified_at, scanned_at) VALUES (?, ?, ?, ?, ?, ?)",
                        batch,
                    )
                    conn.commit()
                    count += len(batch)
                _scan_status["total"] = count
                logger.info(
                    "path_registry_scanned",
                    extra={"count": count, "folders": len(all_folders), "watched": len(watched)},
                )
                update_job_progress(
                    conn,
                    scan_job_id,
                    total_items=len(all_folders),
                    completed_items=len(all_folders),
                    details={
                        "watched_folders": len(watched),
                        "paths_scanned": count,
                        "to_index": len(to_index),
                    },
                    current_stage="walking" if not to_index else "auto_indexing",
                )

                # Phase 2: auto-index text docs in watched folders only
                if to_index:
                    update_job_stage(conn, scan_job_id, "auto_indexing")
                    _auto_index_docs(conn, to_index, parent_job_id=scan_job_id)
                mark_job_completed(conn, scan_job_id, current_stage="completed")

            except Exception as exc:
                mark_job_failed(conn, scan_job_id, str(exc), current_stage="failed")
                logger.exception("path_registry_scan_failed")
            finally:
                _scan_status["running"] = False

            time.sleep(_RESCAN_INTERVAL)

    threading.Thread(target=_scan_loop, daemon=True, name="path-scanner").start()


def _auto_index_docs(conn, file_paths: list[str], parent_job_id: int | None = None):
    """Auto-index text-extractable documents. Skips already-indexed and scanned PDFs."""
    indexed = 0
    failed = 0
    total_files = len(file_paths)
    if parent_job_id is not None:
        update_job_progress(
            conn,
            parent_job_id,
            total_items=total_files,
            completed_items=0,
            details={"indexed": 0, "failed": 0, "total_candidates": total_files},
            current_stage="indexing",
        )
    for processed_count, fpath in enumerate(file_paths, start=1):
        job_id, created = get_or_create_job(conn, "watch_auto_index_file", target_path=fpath, current_stage="queued")
        if not created:
            if parent_job_id is not None:
                update_job_progress(
                    conn,
                    parent_job_id,
                    total_items=total_files,
                    completed_items=processed_count,
                    details={"indexed": indexed, "failed": failed, "total_candidates": total_files},
                    current_stage="indexing",
                )
            continue
        try:
            mark_job_running(conn, job_id, current_stage="indexing")
            update_job_progress(
                conn,
                job_id,
                total_items=1,
                completed_items=0,
                details={"path": fpath},
                current_stage="indexing",
            )
            outcome = index_single_file(
                conn,
                fpath,
                skip_unchanged=True,
                skip_scanned_pdf=True,
                facts_enabled=False,
            )
            if outcome["status"] != "indexed":
                update_job_progress(
                    conn,
                    job_id,
                    total_items=1,
                    completed_items=1,
                    details={"path": fpath, "status": outcome["status"], "reason": outcome.get("reason", "")},
                    current_stage=f"skipped:{outcome.get('reason', 'unknown')}",
                )
                mark_job_completed(conn, job_id, current_stage=f"skipped:{outcome.get('reason', 'unknown')}")
                if parent_job_id is not None:
                    update_job_progress(
                        conn,
                        parent_job_id,
                        total_items=total_files,
                        completed_items=processed_count,
                        details={"indexed": indexed, "failed": failed, "total_candidates": total_files},
                        current_stage="indexing",
                    )
                continue
            indexed += 1
            _scan_status["indexed"] = indexed
            update_job_progress(
                conn,
                job_id,
                total_items=1,
                completed_items=1,
                details={"path": fpath, "status": "indexed"},
                current_stage="completed",
            )
            mark_job_completed(conn, job_id, current_stage="completed")
        except Exception as exc:
            failed += 1
            mark_job_failed(conn, job_id, str(exc), current_stage="failed")
            logger.exception("auto_index_failed", extra={"path": fpath})
        if parent_job_id is not None:
            update_job_progress(
                conn,
                parent_job_id,
                total_items=total_files,
                completed_items=processed_count,
                details={"indexed": indexed, "failed": failed, "total_candidates": total_files},
                current_stage="indexing",
            )

    if indexed:
        logger.info("auto_index_completed", extra={"indexed": indexed})
    if failed:
        logger.warning("auto_index_partial_failures", extra={"failed": failed, "indexed": indexed})


@router.get("/find-file")
def find_file_endpoint(
    query: str = Query(..., description="Filename to search for (case-insensitive)"),
    ext: str = Query("", description="Filter by extension: .pdf, .xlsx, .docx"),
    limit: int = Query(50, ge=1, le=200),
) -> dict:
    """Search the file path registry by filename. Instant — no folder scanning needed."""
    conn = _get_conn()
    conditions = ["filename LIKE ?"]
    params: list = [f"%{query}%"]

    if ext:
        conditions.append("ext = ?")
        params.append(ext.lower() if ext.startswith(".") else f".{ext.lower()}")

    where = " AND ".join(conditions)
    rows = conn.execute(
        f"SELECT path, filename, ext, size_bytes, modified_at FROM file_paths WHERE {where} ORDER BY modified_at DESC LIMIT ?",
        params + [limit],
    ).fetchall()

    results = []
    for r in rows:
        results.append({
            "path": r[0],
            "filename": r[1],
            "ext": r[2],
            "size": r[3],
            "modified": datetime.fromtimestamp(r[4]).strftime("%d %b %Y, %I:%M %p") if r[4] else "",
            "exists": os.path.exists(r[0]),
        })

    return {
        "results": results,
        "count": len(results),
        "_hint": f"{len(results)} file(s) found matching '{query}'." if results else f"No files matching '{query}' in path registry. Try list_files with recursive=true on a specific folder.",
    }


@router.get("/path-scan-status")
def path_scan_status_endpoint() -> dict:
    """Check the status of the background path scan."""
    return {
        "running": _scan_status["running"],
        "paths_scanned": _scan_status["total"],
        "docs_indexed": _scan_status["indexed"],
        "folders": _scan_status["folders"],
    }


class WatchFolderRequest(BaseModel):
    path: str


@router.post("/watch-folder")
def watch_folder_endpoint(req: WatchFolderRequest) -> dict:
    """Start watching a folder — auto-indexes all docs inside, keeps watching for changes."""
    folder = os.path.abspath(req.path)
    if not os.path.isdir(folder):
        raise HTTPException(status_code=404, detail=f"Folder not found: {folder}")
    _check_safe(folder)

    conn = _get_conn()
    existing = conn.execute("SELECT path FROM watched_folders WHERE path = ?", (folder,)).fetchone()
    if existing:
        return {"status": "already_watching", "path": folder, "_hint": f"Already watching {folder}."}

    conn.execute(
        "INSERT INTO watched_folders (path, added_at) VALUES (?, ?)",
        (folder, datetime.now().isoformat()),
    )
    conn.commit()
    watch_job_id = create_job(conn, "watch_initial_index", target_path=folder, current_stage="queued")
    update_job_progress(
        conn,
        watch_job_id,
        total_items=0,
        completed_items=0,
        details={"folder": folder, "discovered": 0},
        current_stage="queued",
    )

    # Kick off immediate indexing in background
    def _index_watched():
        try:
            from database import DB_PATH, init_db

            c = init_db(DB_PATH)
            try:
                mark_job_running(c, watch_job_id, current_stage="walking")
                still_watched = c.execute("SELECT 1 FROM watched_folders WHERE path = ?", (folder,)).fetchone()
                if not still_watched:
                    mark_job_cancelled(c, watch_job_id)
                    return
                to_index = []
                for fpath, fname, ext, size, mtime in _walk_folder(folder):
                    if is_job_cancelling(c, watch_job_id):
                        mark_job_cancelled(c, watch_job_id)
                        return
                    still_watched = c.execute("SELECT 1 FROM watched_folders WHERE path = ?", (folder,)).fetchone()
                    if not still_watched:
                        logger.info("watch_initial_index_cancelled", extra={"folder": folder})
                        mark_job_cancelled(c, watch_job_id)
                        return
                    if ext in _AUTO_INDEX_EXTS and size <= _MAX_AUTO_INDEX_SIZE:
                        to_index.append(fpath)
                update_job_progress(
                    c,
                    watch_job_id,
                    total_items=len(to_index),
                    completed_items=0,
                    details={"folder": folder, "discovered": len(to_index)},
                    current_stage="indexing" if to_index else "completed",
                )
                if to_index:
                    update_job_stage(c, watch_job_id, "indexing")
                    _auto_index_docs(c, to_index, parent_job_id=watch_job_id)
                    logger.info("watch_initial_index_completed", extra={"folder": folder, "count": len(to_index)})
                mark_job_completed(c, watch_job_id, current_stage="completed")
            finally:
                c.close()
        except Exception as exc:
            fail_conn = init_db(DB_PATH)
            try:
                mark_job_failed(fail_conn, watch_job_id, str(exc), current_stage="failed")
            finally:
                fail_conn.close()
            logger.exception("watch_initial_index_failed", extra={"folder": folder})

    threading.Thread(target=_index_watched, daemon=True, name=f"watch-{os.path.basename(folder)}").start()

    return {"status": "watching", "path": folder, "_hint": f"Now watching {folder}. All documents will be indexed + embedded. New/changed files picked up every 60 min."}


@router.post("/unwatch-folder")
def unwatch_folder_endpoint(req: WatchFolderRequest) -> dict:
    """Stop watching a folder."""
    folder = os.path.abspath(req.path)
    _check_safe(folder)
    conn = _get_conn()
    result = conn.execute("DELETE FROM watched_folders WHERE path = ?", (folder,))
    conn.commit()
    cancel_jobs_for_target(conn, "watch_initial_index", folder)
    if result.rowcount > 0:
        return {"status": "unwatched", "path": folder, "_hint": f"Stopped watching {folder}. Existing index data is kept."}
    return {"status": "not_found", "path": folder, "_hint": f"{folder} was not being watched."}


@router.get("/watched-folders")
def watched_folders_endpoint() -> dict:
    """List all watched folders."""
    conn = _get_conn()
    rows = conn.execute("SELECT path, added_at FROM watched_folders ORDER BY added_at").fetchall()
    folders = [{"path": r["path"], "added_at": r["added_at"], "exists": os.path.isdir(r["path"])} for r in rows]
    return {"folders": folders, "count": len(folders)}


@router.get("/background-jobs")
def background_jobs_endpoint(
    status: str = Query("", description="Optional status filter: pending, running, cancelling, completed, failed, cancelled"),
    limit: int = Query(50, ge=1, le=200),
) -> dict:
    conn = _get_conn()
    jobs = list_jobs(conn, limit=limit, status=status)
    return {"jobs": jobs, "count": len(jobs)}


@router.post("/background-jobs/{job_id}/cancel")
def cancel_background_job_endpoint(job_id: int) -> dict:
    conn = _get_conn()
    cancelled = request_job_cancel(conn, job_id)
    if not cancelled:
        raise HTTPException(status_code=404, detail=f"Job not cancellable or not found: {job_id}")
    return {"status": "cancelling", "job_id": job_id}
